import openpyxl
import matplotlib.pyplot as plt
import numpy as np
#
#
# filePath = './documents/angle_error-0.54.xlsx'
# filePath1 = '../documents/0428_sub_edges_5.xlsx'
# filePath2 = '../documents/0428_sub_edges_3.xlsx'
filePath = '../documents/0625_angle_error.xlsx'
# # 打开 XLSX 文件
# wb1 = openpyxl.load_workbook(filePath1)
# wb2 = openpyxl.load_workbook(filePath2)
wb = openpyxl.load_workbook(filePath)
#
# # 选择工作表
# sheet1 = wb1['Sheet1']
# sheet2 = wb2['Sheet1']
sheet = wb['Sheet1']

cal_angle = []
det_angle = []
det_single_angle = []


for j in range(2, 27, 2):
    b = sheet.cell(row=j, column=12)
    c = sheet.cell(row=j, column=14)
    e = sheet.cell(row=j, column=17)

    angle_err1 = abs(c.value-b.value)
    angle_err2 = abs(e.value-b.value)
    if angle_err1<2:
        det_angle.append(angle_err1)
    if angle_err2<2:
        det_single_angle.append(angle_err2)





# for j in range(2, sheet2.max_row + 1):
#     h = sheet2.cell(row=j, column=8)
#     i = sheet2.cell(row=j, column=9)
#     sub_x_3.append(abs(h.value-s_x[j-2]))
#     sub_y_3.append(abs(i.value-s_y[j-2]))
# print(p_y)
# print(np.max(sub_y_3), np.min(sub_y_3))
# print(np.max(sub_x_3), np.min(sub_x_3))
print(np.mean(det_angle), np.mean(det_single_angle))
plt.plot(range(len(det_angle)), det_angle, ">g:")
plt.axhline(np.mean(det_angle), c="g", ls="--")
plt.plot(range(len(det_single_angle)), det_single_angle, "ob-.")
plt.axhline(np.mean(det_single_angle), c="b", ls="--")


plt.show()
