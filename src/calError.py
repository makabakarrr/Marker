import openpyxl
import matplotlib.pyplot as plt
import numpy as np
#
#
# filePath = './documents/angle_error-0.54.xlsx'
# filePath1 = '../documents/0428_sub_edges_5.xlsx'
# filePath2 = '../documents/0428_sub_edges_3.xlsx'
filePath = '../documents/mass/0620_centers_average.xlsx'
# # 打开 XLSX 文件
# wb1 = openpyxl.load_workbook(filePath1)
# wb2 = openpyxl.load_workbook(filePath2)
wb = openpyxl.load_workbook(filePath)
#
# # 选择工作表
# sheet1 = wb1['Sheet1']
# sheet2 = wb2['Sheet1']
sheet = wb['Sheet1']

p_x, p_y = [], []
s_x, s_y = [], []
p_err_x, p_err_y = [], []
sub_x_5, sub_y_5 = [], []
sub_x_3, sub_y_3 = [], []
ransac_1_x, ransac_1_y = [], []
for j in range(2, sheet.max_row + 1):
    b = sheet.cell(row=j, column=2)
    c = sheet.cell(row=j, column=3)
    e = sheet.cell(row=j, column=5)
    f = sheet.cell(row=j, column=6)
    h = sheet.cell(row=j, column=8)
    i = sheet.cell(row=j, column=9)
    k = sheet.cell(row=j, column=11)
    l = sheet.cell(row=j, column=12)
    # n = sheet.cell(row=j, column=14)
    # o = sheet.cell(row=j, column=15)
    p_x.append(e.value)
    p_y.append(f.value)
    s_x.append(b.value)
    s_y.append(c.value)
    p_err_x.append(abs(e.value-b.value))
    p_err_y.append(abs(f.value-c.value))
    sub_x_5.append(abs(h.value - b.value))
    sub_y_5.append(abs(i.value - c.value))
    sub_x_3.append(abs(k.value - b.value))
    sub_y_3.append(abs(l.value - c.value))
    # ransac_1_x.append(abs(n.value - b.value))
    # ransac_1_y.append(abs(o.value - c.value))


# for j in range(2, sheet2.max_row + 1):
#     h = sheet2.cell(row=j, column=8)
#     i = sheet2.cell(row=j, column=9)
#     sub_x_3.append(abs(h.value-s_x[j-2]))
#     sub_y_3.append(abs(i.value-s_y[j-2]))
# print(p_y)
# print(np.max(sub_y_3), np.min(sub_y_3))
# print(np.max(sub_x_3), np.min(sub_x_3))
p_err_x_mean = np.mean(p_err_x)
p_err_y_mean = np.mean(p_err_y)
sub_x_3_mean = np.mean(sub_x_3)
sub_y_3_mean = np.mean(sub_y_3)
plt.subplot(121)
plt.plot(range(len(p_err_x)), p_err_x, color="#038355", marker="o",markeredgecolor="white", linestyle="-", label="pixel")
plt.axhline(p_err_x_mean,color="#038355", linestyle="--")
plt.plot(range(len(sub_x_3)), sub_x_3, color="#ffc34e", marker="s",markeredgecolor="white", linestyle="-", label="sub_pixel")
plt.axhline(sub_x_3_mean, color="#ffc34e", linestyle="--")
plt.ylabel('X-error')
plt.xlabel('Marker')
plt.legend()
# plt.text(2,p_err_x_mean+0.1, str(p_err_x_mean), family="Times New Roman", fontsize=15, style="italic", color="k")
# plt.plot(range(len(sub_x_5)), sub_x_5, "ob-.")
# plt.axhline(np.mean(sub_x_5), c="b", ls="--")

# plt.text(2,sub_x_3_mean, str(sub_x_3_mean), family="Times New Roman", fontsize=15, style="italic", color="k")
# plt.plot(range(len(ransac_1_x)), ransac_1_x, "sk--")
plt.subplot(122)
plt.plot(range(len(p_err_y)), p_err_y, color="#038355", marker="o",markeredgecolor="white", linestyle="-", label="pixel")
plt.axhline(np.mean(p_err_y),color="#038355", linestyle="--")
plt.plot(range(len(sub_y_3)), sub_y_3, color="#ffc34e", marker="s",markeredgecolor="white", linestyle="-", label="sub_pixel")
plt.axhline(np.mean(sub_y_3), color="#ffc34e", linestyle="--")
plt.ylabel('Y-error')
plt.xlabel('Marker')
plt.legend()
# plt.plot(range(len(sub_y_5)), sub_y_5, "ob-.")
# plt.axhline(np.mean(sub_y_5), c="b", ls="--")

# plt.plot(range(len(ransac_1_y)), ransac_1_y, "sk--")
plt.show()
print(p_err_x_mean, sub_x_3_mean)
print(p_err_y_mean, sub_y_3_mean)
