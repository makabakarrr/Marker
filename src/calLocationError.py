import cv2
import openpyxl
import numpy as np
import itertools

from customFunction import calculateAffineMatrix, cvtCodePoints1, calculatePerspectiveMatrix, cvtCodePoints, getRemainPoint

def getDistCombinations(source_combinations, source_points, dist_points):
    src_n0, src_n1, src_n2, src_n3 = source_points
    n0, n1, n2, n3 = dist_points
    dist_combinations  = []
    for s in source_combinations:
        d = []
        for point in s:
            xx, yy = point
            if [xx, yy] == [src_n0[0], src_n0[1]]:
                d.append(n0)
            elif [xx, yy] == [src_n1[0], src_n1[1]]:
                d.append(n1)
            elif [xx, yy] == [src_n2[0], src_n2[1]]:
                d.append(n2)
            else:
                d.append(n3)
        dist_combinations.append(d)

    return np.array(dist_combinations)


filePath = "../documents/0627_location.xlsx"
wb = openpyxl.load_workbook(filePath)
sheet = wb['Sheet1']


source_points = np.array([[10,34], [30, 34], [30, 10], [10, 10]], dtype=np.float32)
target_point = [[22, 22]]
target_real = []
dist_points = []

for row in range(57, 62):
    d = sheet.cell(row=row, column=4)
    e = sheet.cell(row=row, column=5)
    f = sheet.cell(row=row, column=6)
    g = sheet.cell(row=row, column=7)
    sheet.cell(row=row, column=8, value=round(d.value+f.value, 6))
    sheet.cell(row=row, column=9, value=round(e.value+g.value, 6))
    if row<61:
        dist_points.append([round(d.value+f.value, 6), round(e.value+g.value, 6)])
    else:
        target_real.append([round(d.value+f.value, 6), round(e.value+g.value, 6)])
#
# wb.save(filePath)
source_combinations = np.array(list(itertools.combinations(source_points, 3)))
dist_combinations = getDistCombinations(source_combinations, source_points, np.array(dist_points, dtype=np.float32))

M = np.zeros((2,3), np.float32)
for i in range(len(source_combinations)):
    MM = calculateAffineMatrix(source_combinations[i], dist_combinations[i])
    M = np.add(M, MM)
    target_cal_point = cvtCodePoints1(target_point, MM)
    print(i, target_cal_point, (target_cal_point[0][0]-target_real[0][0], target_cal_point[0][1]-target_real[0][1]))

average_matrix = M / len(source_combinations)
average_target = cvtCodePoints1(target_point, average_matrix)
print("average_target:", average_target, (average_target[0][0]-target_real[0][0], average_target[0][1]-target_real[0][1]))

# source_combinations = np.array(list(itertools.combinations(source_points, 3)))
# dist_combinations = getDistCombinations(source_combinations, source_points, np.array(dist_points, dtype=np.float32))
# angle = 0.0
# transform_matrix = np.zeros((2,3), np.float32)
# max_dist = 10.0

# for i in range(len(source_combinations)):
#     MM = calculateAffineMatrix(source_combinations[i], dist_combinations[i])
#     s_remain_p, d_remain_p = getRemainPoint(source_combinations[i], source_points, dist_points)
#     cal_dist_x = round(MM[0][0] * s_remain_p[0] + MM[0][1] * s_remain_p[1] + MM[0][2], 4)
#     cal_dist_y = round(MM[1][0] * s_remain_p[0] + MM[1][1] * s_remain_p[1] + MM[1][2], 4)
#     dist_err = np.sqrt((cal_dist_x-d_remain_p[0])**2 + (cal_dist_y-d_remain_p[1])**2)
#     print(dist_err, 135 - np.arctan2(MM[1, 0], MM[0, 0]) * 180 / np.pi)
#     if dist_err<max_dist:
#         max_dist = dist_err
#         angle = 135 - np.arctan2(MM[1, 0], MM[0, 0]) * 180 / np.pi
#         print("min:", dist_err, angle)
#         transform_matrix = MM
#         print("remain_point:", d_remain_p, cal_dist_x, cal_dist_y, (cal_dist_x-d_remain_p[0], cal_dist_y-d_remain_p[1]))
# angle = angle if angle>0 else 360+angle
# print("标识的角度为：", angle)
# average_target = cvtCodePoints1(target_point, transform_matrix)
# print("计算值：", average_target, (average_target[0][0]-target_real[0][0], average_target[0][1]-target_real[0][1]))