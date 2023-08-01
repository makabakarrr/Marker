import random
import numpy as np
import matplotlib.pyplot as plt
import openpyxl

plt.rcParams['font.sans-serif']=['kaiti']

m = 1
R = 96


def six_variable_function(params):
    """
    求I20的函数值
    :param xx1: 需要求的参数
    :return: I20的函数值
    """
    y = 0.0
    x1, x2, x3, x4, x5 = 86.3, 83.9, 0, 3.54, 0.00125    # x1--rA x2--rH x3--α  x4--βH
    x6 = params[0] # 若params里只有1个值
    # x6, x7 = [np.float(xx) for xx in params] # 若params里只有2个值,有多个值以此类推
    x3, x4 = np.radians(x3), np.radians(x4)     # 把角度转为弧度值,因为后续的三角函数的参数要求为弧度制
    for l in range(350, 851):   # λ的取值范围350-850，步长为1
        l = l*0.000001   #nm----mm
        temp = np.clip(m*l/x5 - np.sin(x3), -1, 1)  # m*l/x5 - np.sin(x3)---np.clip的作用是将其限制在-1到1的范围内，arcsin的定义域
        y1 = np.arcsin(temp)  # 角度的单位为弧度---β(λq)
        y2 = x2 / (np.cos(x4-y1))
        y3 = np.cos(x3)**2 / x1 + np.cos(y1)**2 / y2 - (np.cos(x3)+np.cos(y1)) / R
        y4 = (y3 + m*l/x5 * x6)**2
        y += y4 # y代表I20

    return y



# 定义遗传算法参数
population_size = 70   # 种群数量
max_generations = 100    # 迭代代数
mutation_rate = 0.5     # 变异率------决定个体是否发生变异
elite_size = 10          # 精英数量------每次的种群里保留10个精英个体


# 定义变量范围和编码长度
# variable_ranges = [(0, 200), (0, 200), (-90, 90), (-90, 90), (0.5, 5), (-10, 10)]
# variable_length = [20, 20, 15, 15, 16, 18]
variable_ranges = [(-10, 10)]   # 参数的取值范围(a, b),代表参数可从a取到b，若有其他参数，以此形式添加[(a1,b1), (a2,b2), (a3, b3)...]
variable_length = [18]  # 二进制串的长度n   计算方式(b-a)/((2^n)-1)≤0.0001---0.0001为参数的精度 若有其他参数，以此形式添加[n1, n2, n3] variable_length[i]对应variable_ranges[i]





def generate_population():
    """
    生成初始种群
    :return: 初始种群
    """
    population = []
    for _ in range(population_size):
        individual = []
        for i in range(len(variable_ranges)):
            individual.append(''.join(random.choice('01') for _ in range(variable_length[i])))  # 随机生成二进制串
        population.append(individual)
    return population


def decoded_x(individual):
    """
    对二进制串进行解码
    :param individual: 参数列表对应二进制串列表
    :return: 解码后的参数列表
    """

    params_list = [variable_ranges[i][0] + int(individual[i], 2) * (variable_ranges[i][1] - variable_ranges[i][0]) / (
                np.power(2, variable_length[i]) - 1) for i in range(len(individual))] # 解码过程：值= a+T*p 其中a为参数范围的左边界，T为二进制串对应的十进制，p为参数的精度

    return params_list



def fitness_function(individual):
    """
    计算适应度函数：要求的是I20的极小值，选择时选择适应度小的作为精英个体
    :param individual: 个体---参数列表对应的二进制串列表
    :return: 适应度---此处代表I20
    """
    params = decoded_x(individual)

    print("函数值：", six_variable_function(params))

    return  six_variable_function(params)


# 选择操作
def selection(population):
    """
    根据适应度排序选择部分个体（适应度最小）
    :param population: 种群
    :return: 择优选取后的精英种群
    """
    sorted_population = sorted(population, key=fitness_function)    # 将population按适应度从小到大的顺序排列
    return sorted_population[:elite_size]   # 选择前elite_size个精英个体


# 交叉操作
def crossover(parent1, parent2):
    """
    使用两个个体交叉生成新个体
    :param parent1: 个体1
    :param parent2: 个体2
    :return: 新个体
    """
    child = []
    list1, list2 = list(parent1[0]), list(parent2[0])   # 将二进制字符串转为数组，python中不允许对字符串进行修改
    point1 = random.randint(0, variable_length[0]-1)    # 在0到variable_length[0]-1范围内随机取一个整数
    point2 = random.randint(0, variable_length[0]-1)
    if point1 > point2:
        point1, point2 = point2, point1
    new_ = list1[:point1] + list2[point1:point2] + list1[point2:]
    child.append("".join(new_))
    return child


# 变异操作
def mutate(chromosome):
    """
    对个体的染色体进行变异，只对二进制字符串变异了一个字符----若这字符是1则变为0，若为0则变为1
    :param chromosome: 个体
    :return: 变异后的个体
    """
    for i in range(len(variable_ranges)):
        if random.random() < mutation_rate: # 根据变异率进行变异
            chromosome_list = list(chromosome[i])
            mutation_point = random.randint(0, variable_length[i] - 1)
            chromosome_list[mutation_point] = '1' if chromosome_list[mutation_point] == '0' else '0'
            chromosome[i] = ''.join(chromosome_list)
    return chromosome


# 遗传算法主流程
def genetic_algorithm():
    best_list = []  # 存放每次迭代的最佳适应度
    population = generate_population()  # 初始化种群

    for j in range(max_generations):
        # 择优选择---选择精英作为遗传对象
        new_population = selection(population)
        if j!=0:
            best_x = new_population[0]
            best_fitness = fitness_function(best_x)
            best_list.append(best_fitness)

        # 根据精英种群进行交叉和变异，保证种群内的个体数为population_size
        while len(new_population) < population_size:
            parent1 = random.choice(population)
            parent2 = random.choice(population)
            child = crossover(parent1, parent2) # 随机选取两个个体进行交叉
            child = mutate(child)   # 变异
            new_population.append(child)

        population = new_population

    best_individual = selection(population)[0]  # 迭代结束后的最优个体
    return best_individual, best_list


# 运行遗传算法
best_solution, fitness_list = genetic_algorithm()
best_params = decoded_x(best_solution)
best_fitness = six_variable_function(best_params)

# 输出结果
print("Best Params:", (best_params))
print("Best fitness:", best_fitness)
# filePath = '../documents/genetic.xlsx'
# wb = openpyxl.load_workbook(filePath)
#
# sheet = wb['Sheet1']
# row = sheet.max_row + 1
# sheet.cell(row=row, column=1, value=row-1)
# sheet.cell(row=row, column=2, value=round(x1, 4))
# sheet.cell(row=row, column=3, value=round(x2, 4))
# sheet.cell(row=row, column=4, value=round(x3, 4))
# sheet.cell(row=row, column=5, value=round(x4, 4))
# sheet.cell(row=row, column=6, value=round(x5, 4))
# sheet.cell(row=row, column=7, value=round(x6, 4))
# wb.save(filePath)

plt.plot(fitness_list)
plt.title("适应度曲线")
# plt.savefig("../images/genetic/4.png")
plt.show()







