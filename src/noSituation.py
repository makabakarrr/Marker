# import cv2
# import numpy as np
import time

import cv2
import openpyxl
import matplotlib.pyplot as plt
import sys
import itertools

import numpy as np

from customFunction import *
from utils import calculateCrossPoint, getAngle, calculateAngle
from ransacFunction import ransac_circle_fit

from sklearn.cluster import DBSCAN



def cal_gaussian(Z, X):
    A = np.zeros((len(X), 3))
    for i in range(3):
        A[:, i] = np.power(X, i)
    A_T = A.T
    Z_T = Z.T
    ATA = np.dot(A_T, A)
    b = np.dot(A_T, Z_T)
    if np.linalg.det(ATA)==0:
        return 0
    else:
        B = np.dot(np.linalg.inv(ATA), b)

        BB = -1*B[1]/(2*B[2])
        if BB>3:
            BB = 3
        elif BB<0.0001:
            BB = 0.0001
        return 0 if np.isnan(BB) else BB

def gaussian_fit(p):
    x, y = p

    dx = abs(sobelx)
    dy = abs(sobely)
    # theta = sobely[y][x] / sobelx[y][x]
    # print(np.min(angle), np.max(angle))
    theta = np.degrees(sobel_angle[y][x])
    step = 180 / 8
    if -1*step <= theta < theta:
        y_bar = [0] * 5
        x_bar = [x for x in range(-2, 3)]
    elif step <= theta < 3*step:
        y_bar = [y for y in range(2, - 3, -1)]
        x_bar = [x for x in range(-2, 3)]
    elif 3*step<=theta or theta <= -3*step:
        x_bar = [0] * 5
        y_bar = [y for y in range(2, -3, -1)]
    else:
        y_bar = [y for y in range(-2, 3)]
        x_bar = [x for x in range(-2, 3)]
    # if -1*step <= theta < theta:
    #     y_bar = [0] * 3
    #     x_bar = [x for x in range(-1, 2)]
    # elif step <= theta < 3*step:
    #     y_bar = [y for y in range(1, - 2, -1)]
    #     x_bar = [x for x in range(-1, 2)]
    # elif 3*step<=theta or theta <= -3*step:
    #     x_bar = [0] * 3
    #     y_bar = [y for y in range(1, -2, -1)]
    # else:
    #     y_bar = [y for y in range(-1, 2)]
    #     x_bar = [x for x in range(-1, 2)]

    # 获取五组数据的梯度幅值
    mag_y = [dy[j+y][i+x] for j,i in list(zip(y_bar, x_bar))]
    mag_x = [dx[j+y][i+x] for j,i in list(zip(y_bar, x_bar))]

    # 排序
    # print("mag:", mag_x, mag_y)
    subx = cal_gaussian(np.array(mag_x, dtype=np.float32), x_bar)

    suby = cal_gaussian(np.array(mag_y, dtype=np.float32), y_bar)

    return x+subx, y+suby


def getSubEdgePoints(circle_points):
    cur_sub_edges = []
    for p in circle_points:
        xx, yy = gaussian_fit(p)
        # cur_sub_edges.append([xx, yy])
        cur_sub_edges.append([[xx, yy]])

    return cur_sub_edges


def extractCircles(cnts, no_branch, edges_points):
    circle_canvas = cv2.cvtColor(no_branch, cv2.COLOR_GRAY2BGR)
    circle_info = []
    circle_edge_points = []
    for cnt_index in range(len(cnts)):
        cnt = cnts[cnt_index]
        if len(cnt) > 10:
            (cx, cy), (a, b), angle = cv2.fitEllipse(cnt)
            radius = max(a, b) / 2
            # cv2.circle(circle_canvas, (int(cx), int(cy)), int(radius), (0, 0, 255), 1)
            if min(a,b)<20 or max(a, b)>200:
                continue
            # cv2.circle(circle_canvas, (int(cx), int(cy)), int(radius), (255, 0, 0), 1)
            # 求内点数量
            perimeter = 2 * np.pi * radius
            nums = 0
            test_points = []
            for point in cnt:
                x, y = point[0]
                if [x, y] not in test_points:
                    test_points.append([x, y])

            for point in test_points:
                x, y = point
                dist = abs(np.sqrt((x - cx) ** 2 + (y - cy) ** 2) - radius)
                if dist < 2:  # test_point 避免重复
                    nums += 1

            k = nums / perimeter
            # cv2.circle(circle_canvas, (int(cx), int(cy)), int(radius), (0, 0, 255), 1)
            # cv2.putText(circle_canvas, str(radius), (int(cx)+int(radius), int(cy)), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,255,0), 1)
            cv2.circle(circle_canvas, (int(cx), int(cy)), int(radius), (255, 0, 0), 1)
            if k > 0.3:  # 圆形轮廓
                circle_info.append([cx, cy, radius, len(circle_edge_points), k])    # [中心坐标x, 中心坐标y, 对应的边缘点索引号, k值]
                cv2.circle(circle_canvas, (int(cx), int(cy)), int(radius), (0, 0, 255), 1)
                for p in test_points:
                    i, j = p
                    filter_cnts[j][i] = 255
                circle_edge_points.append(test_points)


    cv2.imwrite("../images/process/summary/filterCnts/" + imgName + '-filter_cnts.png', filter_cnts)
    cv2.imwrite("../images/process/summary/filterCnts/" + imgName + '-fit_circle.png', circle_canvas)
    return circle_info, circle_edge_points


def classifyArcByPosition(circle_info):
    circle_centers = np.array(circle_info)[:, 0:2]
    db = DBSCAN(eps=20, min_samples=1).fit(circle_centers)
    c_s = db.labels_
    # plt.imshow(img, cmap="gray")
    # plt.scatter(np.array(circle_info)[:, 0], np.array(circle_info)[:, 1], s=10, marker="*", c=db.labels_)
    # plt.show()
    return c_s

#
# def classifyArcAndFilter(circle_info, circle_edge_points, category_index):
#     start_template_circle_points = []  # 起始模板点的边缘点 [[起始模板点1 [圆形1 [点1], [点2]], [圆形2]], [起始模板点2]]
#     start_template_circle_info = []
#     other_template_circle_points = []  # 其他模板点的边缘点
#     other_template_circle_info = []
#     locate_circle_points = []   # 定位圆的边缘点
#     locate_circle_info = []
#     code_circle_points = []  # 编码圆
#     code_circle_info = []
#     # 获取最大圆半径 max_radius
#     all_radius = np.array(circle_info)[:, 2]
#     sorted_all_radius = np.argsort(all_radius)
#     max_radius = circle_info[sorted_all_radius[-1]][2]
#     print("DBSCAN将轮廓分成了：", np.max(category_index) + 1, "类")
#     for c in range(np.max(category_index) + 1):
#         cate_indexs = np.where(category_index == c)[0]  # 某一类的索引
#         cate_circles = np.array(circle_info)[cate_indexs]   #属于该类里的所有圆信息
#         mean_x, mean_y = np.median(cate_circles[:,0]), np.median(cate_circles[:, 1])    #获取该类圆心集合的中心---为了滤除异常值
#         print("集合中心点：", mean_x, mean_y)
#         cur_cate_radius = cate_circles[:, 2]
#         sorted_radius = np.argsort(cur_cate_radius)
#         sorted_circles = cate_circles[sorted_radius]  # 对该类的圆参数按半径从小到大进行排序
#         radius_flag = 0  # 最小的半径值-----根据半径对同一类别里的圆进行细分
#         nums_flag = 0   # 内点数量/拟合的圆的周长---边缘点越完整值越大
#         radius_category = 1 # 统计当前类别里有几种半径，
#         all_edges_points = []   # 某一大类内的所有圆形边缘点
#         single_circle_points = []   # 某个小类别里的所有圆形边缘点
#         for circle_index in range(len(sorted_circles)):
#             circle = sorted_circles[circle_index]
#             if np.sqrt((circle[0]-mean_x)**2 + (circle[1]-mean_y)**2)>2:
#                 print("超出范围", circle)
#                 continue
#             if len(single_circle_points) == 0:
#                 radius_flag = circle[2]
#                 nums_flag = circle[4]
#                 single_circle_points = [[p[0][0], p[0][1]] for p in circle_edge_points[int(circle[3])]]
#             else:
#                 if circle[2]-radius_flag < 7:  # 同一个圆形
#                     for p in circle_edge_points[int(circle[3])]:
#                         i, j = p[0]
#                         if [i, j] not in single_circle_points:
#                             single_circle_points.append([i, j])
#                     if nums_flag<circle[4]:
#                         nums_flag = circle[4]
#                         radius_flag = circle[2]
#                 else:
#                     radius_flag = circle[2]
#                     nums_flag = circle[4]
#                     radius_category += 1
#                     all_edges_points.append(single_circle_points)
#                     single_circle_points = [[p[0][0], p[0][1]] for p in circle_edge_points[int(circle[3])]]
#         all_edges_points.append(single_circle_points)
#         # print("半径种类：", radius_category)
#         # print(sorted_circles)
#         if radius_category == 1:  # 1种半径  定位圆和编码圆
#             if max_radius-radius_flag < 5:   # 只有一种半径值，并且是标记点中半径最大的圆，标记点中半径最大的圆有四个模板点最外层的圆、中心定位圆
#                 locate_circle_points.append(all_edges_points)
#                 locate_circle_info.append(sorted_circles)
#             else:
#                 code_circle_points.append(all_edges_points)
#                 code_circle_info.append(sorted_circles)
#         elif radius_category == 2:  # 2种半径  起始模板点
#             start_template_circle_points.append(all_edges_points)
#             start_template_circle_info.append(sorted_circles)
#         else:  # 其他类模板点
#             other_template_circle_points.append(all_edges_points)
#             other_template_circle_info.append(sorted_circles)
#     return start_template_circle_points, start_template_circle_info, other_template_circle_points, other_template_circle_info, locate_circle_points, locate_circle_info, other_circle_points, other_circle_info


def classifyArcByRadius(circle_info, category_index):
    all_circle_tree = []
    print("DBSCAN将轮廓分成了：", np.max(category_index) + 1, "类")
    for c in range(np.max(category_index) + 1):
        cate_indexs = np.where(category_index == c)[0]  # 某一类的索引
        cate_circles = np.array(circle_info)[cate_indexs]   #属于该类里的所有圆信息[[circle1_x, circle1_y...], [circle2_x, circle_y,....]]
        # mean_x, mean_y = np.median(cate_circles[:,0]), np.median(cate_circles[:, 1])    #获取该类圆心集合的中心---为了滤除异常值
        max_k_index = np.argmax(cate_circles[:,4])
        mean_x, mean_y = cate_circles[max_k_index][0], cate_circles[max_k_index][1]
        # mean_x, mean_y = np.mean(cate_circles[:,0]), np.mean(cate_circles[:, 1])    #获取该类圆心集合的中心---为了滤除异常值
        # print("集合中心点：", mean_x, mean_y)
        cur_cate_radius = cate_circles[:, 2]    # 属于该类的所有圆的半径：[circle1_r, circle2_r, ...]
        sorted_radius = np.argsort(cur_cate_radius)
        sorted_circles = cate_circles[sorted_radius]  # 对该类的圆参数按半径从小到大进行排序
        # print("sorted_radius:", sorted_circles)
        radius_flag = 0  # 最小的半径值-----根据半径对同一类别里的圆进行细分
        nums_flag = 0   # 内点数量/拟合的圆的周长---边缘点越完整值越大
        radius_category = 1 # 统计当前类别里有几种半径，
        cur_cate_all_circle = []   # 某一大类内的所有圆形边缘点
        cur_sub_cate_circle = []   # 某个小类别里的所有圆形边缘点
        for circle_index in range(len(sorted_circles)):
            circle = sorted_circles[circle_index]
            if np.sqrt((circle[0]-mean_x)**2 + (circle[1]-mean_y)**2)>2:    # 异常圆
                print(mean_x, mean_y, np.sqrt((circle[0]-mean_x)**2 + (circle[1]-mean_y)**2))
                print("异常圆", circle)
                continue
            if len(cur_sub_cate_circle) == 0:
                radius_flag = circle[2]
                nums_flag = circle[4]
                cur_sub_cate_circle = [circle]
            else:
                if circle[2]-radius_flag < 7:  # 同一个圆形
                    cur_sub_cate_circle.append(circle)
                    if nums_flag<circle[4]:
                        nums_flag = circle[4]
                        radius_flag = circle[2]
                else:
                    radius_flag = circle[2]
                    nums_flag = circle[4]
                    radius_category += 1
                    cur_cate_all_circle.append(cur_sub_cate_circle)
                    cur_sub_cate_circle = [circle]
        if cur_sub_cate_circle:
            cur_cate_all_circle.append(cur_sub_cate_circle)
            all_circle_tree.append(cur_cate_all_circle)
    return all_circle_tree


def showClassifyArcByPosition(category_indexs, circle_info, circle_edge_points):
    classify_canvas = np.zeros((height, width, 3), np.uint8)
    # classify_canvas *= 255
    for c in range(np.max(category_indexs) + 1):
        color = (np.random.randint(0, 255), np.random.randint(0, 255), np.random.randint(0, 255))
        cate_indexs = np.where(category_indexs == c)[0]  # 某一类的索引
        for index in cate_indexs:
            circle = circle_info[index]
            circle_edges = circle_edge_points[int(circle[3])]
            for point in circle_edges:
                x,y = point
                classify_canvas[y, x, :] = color
    cv2.imwrite("../images/process/angle/classifyByPos/" + imgName + "_pos_category.png", classify_canvas)


def showClassifyArcByRadius(classify_tree, circle_edge_points, name):
    classify_canvas = np.ones((height, width, 3), np.uint8)
    classify_canvas *= 255
    for position in classify_tree:
        colors = []
        for i in range(len(position)):
            colors.append((np.random.randint(0, 255), np.random.randint(0, 255), np.random.randint(0, 255)))

        for j in range(len(position)):
            radius = position[j]
            for circle in radius:
                cur_circle_points = circle_edge_points[int(circle[3])]
                for point in cur_circle_points:
                    x, y = point
                    classify_canvas[y, x, :] = colors[j]
    cv2.imwrite("../images/process/summary/" + name +"/" + imgName + "_category.png", classify_canvas)

def showFilterArc(classify_tree, circle_edge_points, name):
    classify_canvas = np.ones((height, width, 3), np.uint8)
    classify_canvas *= 255
    for position in classify_tree:
        colors = []
        for i in range(len(position)):
            colors.append((np.random.randint(0, 255), np.random.randint(0, 255), np.random.randint(0, 255)))

        for j in range(len(position)):
            circle = position[j]
            cur_circle_points = circle_edge_points[int(circle[3])]
            for point in cur_circle_points:
                x, y = point
                classify_canvas[y, x, :] = colors[j]
    cv2.imwrite("../images/process/summary/" + name + "/" + imgName + "_category.png", classify_canvas)




def calCenters(circle_Points):
    center_list = [] # 几组圆的中心坐标
    # 有几组圆
    for cate in circle_Points:
        # 每一组圆里有几个圆
        cur_cate_center = []
        for circle in cate:
            # 使用椭圆拟合的方式计算中心坐标点
            (x, y), (a,b), angle = cv2.fitEllipse(np.array(circle))
            cur_cate_center.append([x, y])
        center_mean = np.mean(np.array(cur_cate_center), axis=0)
        center_list.append(np.array(center_mean))
    return center_list


def calCentersByAverage(center_info):
    cate_center= []
    for cate in center_info:
        mean_val = np.array(np.mean(np.array(cate)[:, 0:2], axis=0).round(4))
        cate_center.append([mean_val[0], mean_val[1]])
    return cate_center


def gaussianFit(edge_points):
    for point in edge_points:
        x, y = point.ravel()

        # 获取3x3邻域图像块
        image_block = img[int(y) - 1:int(y) + 2, int(x) - 1:int(x) + 2]

        # 计算图像块的梯度
        gradient_x = cv2.Sobel(image_block, cv2.CV_64F, 1, 0, ksize=3)
        gradient_y = cv2.Sobel(image_block, cv2.CV_64F, 0, 1, ksize=3)

        # 将梯度转换为向量形式
        gradient_x = gradient_x.flatten()
        gradient_y = gradient_y.flatten()

        # 组合梯度向量
        gradient_vector = np.vstack((gradient_x, gradient_y)).T

        # 进行线性拟合
        line_direction = cv2.fitLine(gradient_vector, cv2.DIST_L2, 0, 0.01, 0.01)

        # 提取拟合参数
        vx, vy, x0, y0 = line_direction

        # 计算亚像素边缘坐标点
        subpixel_x = x0[0] - vx[0] * y0[0] / vy[0]
        subpixel_y = y0[0] - vy[0] * x0[0] / vx[0]



def getDistCombinations(source_combinations, source_points, dist_points):
    src_n0, src_n1, src_n2, src_n3 = source_points
    n0, n1, n2, n3 = dist_points
    dist_combinations  = []
    for s in source_combinations:
        d = []
        for point in s:
            xx, yy = point
            if [xx, yy] == [src_n0[0], src_n0[1]]:
                d.append(n0)
                # d.append('n0')
            elif [xx, yy] == [src_n1[0], src_n1[1]]:
                d.append(n1)
                # d.append('n1')
            elif [xx, yy] == [src_n2[0], src_n2[1]]:
                d.append(n2)
                # d.append('n2')
            else:
                d.append(n3)
                # d.append('n3')
        dist_combinations.append(d)

    return np.array(dist_combinations)


def getRemainPoint(source_combination, source_points, dist_points):

    src_n0, src_n1, src_n2, src_n3 = [[x,y] for x,y in source_points]
    n0, n1, n2, n3 = [[x,y] for x,y in dist_points]
    s_points = [[x, y] for x, y in source_combination]
    source_remain_point = src_n0
    dist_remain_point = n0

    if src_n0 not in s_points:
        print("组合里没有src_n0")
        source_remain_point = src_n0
        dist_remain_point = n0
    elif src_n1 not in s_points:
        print("组合里没有src_n1")
        source_remain_point = src_n1
        dist_remain_point = n1
    elif src_n2 not in s_points:
        print("组合里没有src_n2")
        source_remain_point = src_n2
        dist_remain_point = n2
    else:
        print("组合里没有src_n3")
        source_remain_point = src_n3
        dist_remain_point = n3

    return source_remain_point, dist_remain_point


def delBranch(edges, angle_thresh):
    sub_mask1 = np.zeros((edges.shape[0], edges.shape[1], 3), np.uint8)
    sub_mask2 = np.zeros(edges.shape, np.uint8)
    sub_mask3 = np.zeros(edges.shape, np.uint8)
    cnts0, _ = cv2.findContours(edges, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
    # for cnt in cnts0:
    for index in range(len(cnts0)):
        cnt = cnts0[index]
        if len(cnt) > 10:
            start_x, start_y = cnt[0][0]
            # cv2.putText(sub_mask1, str(index), (start_x, start_y), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,0,255), 1)
            cv2.drawContours(sub_mask1, [cnt], -1, (255,255,255), 1)
            cv2.drawContours(sub_mask2, [cnt], -1, (255,255,255), 1)
            cv2.drawContours(sub_mask3, [cnt], -1, (255,255,255), 1)
            epsilon = 1
            approx = cv2.approxPolyDP(cnt, epsilon, False)
            ### approx可视化
            colors = []
            for i in range(len(approx)):
                colors.append((np.random.randint(0, 255), np.random.randint(0, 255), np.random.randint(0, 255)))
            for ap_index in range(len(approx) - 1):
                i, j = approx[ap_index][0]
                ii, jj = approx[ap_index + 1][0]
                # sub_mask1[j, i, :] = colors[ap_index]
                cv2.line(sub_mask1, [i, j], [ii, jj], colors[ap_index], 1)

            for index in range(1, len(approx) - 1):
                p_0 = np.array(approx[index - 1][0])
                p_1 = np.array(approx[index][0])
                p_2 = np.array(approx[index + 1][0])
                l_0 = p_0 - p_1
                l_1 = p_1 - p_2

                aa = np.dot(l_0, l_1) / (np.linalg.norm(l_0) * np.linalg.norm(l_1))
                aa = 1 if aa>1.0 else aa
                aa = -1 if aa<-1.0 else aa
                A = np.degrees(np.arccos(aa))

                if A > angle_thresh:
                    xx, yy = p_1
                    sub_mask2[yy][xx] = 0
                    # sub_mask1[yy,xx,:] = (0,0,255)

    cv2.imwrite("../images/process/summary/approxResult/" + imgName + "_cnt_approx.png", sub_mask1)
    cv2.imwrite("../images/process/summary/approxResult/" + imgName + "_cnt_break.png", sub_mask2)
    return sub_mask2


def filterArcByMaxK(circle_tree):
    # 每一种半径类别里只保留k值最大的那个圆
    # 则每一个位置内有几个元素就代表该位置上有多少种半径，每一个元素均为该种半径内的k值最大的圆的参数信息[[pos1-[circle_x, circle_y,....]]]
    simplify_tree = []
    for pos_cate in circle_tree:
        pos_max_circle = []
        for radius_cate in pos_cate:
            # print("radius_cate:", radius_cate)
            cur_radius_k = np.array(radius_cate)[:, 4]
            sorted_k = np.argsort(cur_radius_k)
            pos_max_circle.append(radius_cate[sorted_k[-1]])
        simplify_tree.append(pos_max_circle)
    return simplify_tree


def fillIntCircle(no_branch, other_circle_info):
    canvas = no_branch.copy()
    for other_circle in other_circle_info:
        # mean_val = np.array(np.mean(np.array(other_circle)[:, 0:3], axis=0))
        cv2.circle(canvas, (int(other_circle[0]), int(other_circle[1])), int(other_circle[2]), (255,255,255), -1)
    # cv2.imwrite("../images/process/angle/decode/" + imgName + "_coded_circles.png", canvas)
    return canvas


def decodedByGray(points, lp, side_length, img):
    canvas = cv2.cvtColor(edges, cv2.COLOR_GRAY2BGR)
    # 根据灰度值进行解码----灰度值几乎无差别，区分不出来
    step = side_length / 10
    temp_dist = step * 2
    gray = 0.0
    # 随机取四个方向的模板
    for i in range(4):
        angle = random.randint(0, 360)
        x = int(lp[0] + temp_dist*np.cos(np.radians(angle)))
        y = int(lp[1] - temp_dist*np.sin(np.radians(angle)))
        cv2.circle(canvas, (x,y), 2, (0,0,255), -1)
        surrounding = img[y-5:y+6, x-5:x+6]
        gray_mean = np.mean(surrounding)
        gray += gray_mean
    gray /= 4
    int_str = ''
    for index in range(len(points)):
        p = points[index]
        i, j = int(p[0]), int(p[1])
        cv2.putText(canvas, str(index), (i, j), cv2.FONT_HERSHEY_SIMPLEX, 1, (247, 162, 17), 2)
        cv2.circle(canvas, (i, j), 2, (0, 255, 0), -1)
        p_surrounding = img[j-5:j+6, i-5:i+6]
        p_gray = np.mean(p_surrounding)
        print(p_gray, gray)
        if abs(p_gray - gray) > 5:
            int_str += '1'
        else:
            int_str += '0'
    # cv2.imwrite("../images/process/angle/decode/" + imgName + '-pos.png', canvas)
    return int_str


def extractCircularEdges(sobel_thresh, side_length):
    height, width = edges.shape
    dec_edges = np.zeros((height, width), np.uint8)
    dec_rect = np.zeros((height, width, 3), np.uint8)
    circular_cnts = []
    # 提取小数部分的边缘信息
    length = int(side_length / 10 * 3.8)
    locate_radius = int(side_length / 10 * 1.8)
    contours, _ = cv2.findContours(sobel_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    for cnt in contours:
        if len(cnt) > 20:
            # 获取x的最大值和最小值
            x_max = np.max(cnt[:, 0, 0])
            x_min = np.min(cnt[:, 0, 0])
            # 获取y的最大值和最小值
            y_max = np.max(cnt[:, 0, 1])
            y_min = np.min(cnt[:, 0, 1])
            isCircular = False
            if lp[0] - length < x_min and x_max < lp[0] + length and lp[1] - length < y_min and y_max < lp[1] + length:
                isCircular = True

            if lp[0] - locate_radius < x_min and x_max < lp[0] + locate_radius and lp[
                1] - locate_radius < y_min and y_max < lp[1] + locate_radius:
                isCircular = False

            if isCircular:
                circular_cnts.append(cnt)
                cv2.drawContours(dec_edges, [cnt], -1, (255, 255, 255), -1)
                rect = cv2.minAreaRect(cnt)
                box = cv2.boxPoints(rect)
                box = np.int0(box)
                cv2.drawContours(dec_rect, [box], 0, (0, 0, 255), 1)
                x, y, w, h = cv2.boundingRect(cnt)
                cv2.rectangle(dec_rect, (x, y), (x + w, y + h), (0, 255, 0), 1)
                cv2.drawContours(dec_rect, [cnt], -1, (255, 255, 255), 1)
                # (dec_x, dec_y), (dec_a, dec_b), dec_angle = cv2.fitEllipse(cnt)
                # cv2.ellipse(dec_rect, (int(dec_x), int(dec_y)), (int(dec_a/2), int(dec_b/2)), dec_angle, 0, 360, (0,0,255), 1)

    # cv2.imwrite("../images/process/angle/decode/dec_edges/" + imgName + "_dec_edges.png", dec_edges)
    # cv2.imwrite("../images/process/angle/decode/dec_rect/" + imgName + "_dec_ellipse.png", dec_rect)
    return dec_edges


def showExtractArc(circle_edge_points, name):
    arc_edge = np.zeros((height, width, 3), np.uint8)
    arc_edge *= 255
    colors = []
    for i in range(len(circle_edge_points)):
        colors.append((np.random.randint(0, 255), np.random.randint(0, 255), np.random.randint(0, 255)))
    for circle_index in range(len(circle_edge_points)):
        points = circle_edge_points[circle_index]
        for point in points:
            x, y = point
            arc_edge[y, x, :] = colors[circle_index]
    # cv2.imwrite("../images/process/angle/" + name +"/" + imgName + '_arc.png', arc_edge)


def isSameCircle(circle1, circle2, dist_threshold, radius_threshold):
    # 计算两个圆心之间的距离
    dist = math.sqrt((circle1[0]-circle2[0])**2 + (circle1[1]-circle2[1])**2)
    # 计算两个圆半径之差
    radius_diff = abs(circle1[2] - circle2[2])
    # 判断是否代表同一个圆
    if (dist <= dist_threshold) and (radius_diff <= radius_threshold):
        return True
    else:
        return False

def groupCircles(circles, circle_edges_points, dist_threshold, radius_threshold):
    num_circles = len(circles)
    # 初始化将所有圆都归为不同的类别
    # groups = [0] * num_circles
    groups = [[] for i in range(num_circles)]
    # 设定类别id
    group_id = 0
    for i in range(num_circles):
        # 如果该圆已经被分入某一个类别，则跳过
        # if groups[i] != 0:
        #     print("已分类")
        if groups[i]:
            continue
        # 新建一个类别
        groups[i].append(group_id)
        # groups[i] = group_id
        # 将与该圆代表同一个圆的其他圆分入该类别
        for j in range(i+1, num_circles):
            if isSameCircle(circles[i], circles[j], dist_threshold, radius_threshold):
                groups[j].append(group_id)
                # groups[j] = group_id
        group_id += 1
    groups_list = np.array(groups).flatten()
    new_circle_info = []
    new_circle_points = []
    for i in range(0, np.max(groups_list)+1):
        cate_indexs = np.where(groups_list == i)[0]
        cate_circle = []
        # 边缘点去重
        circle_points = []
        for index in cate_indexs:
            cate_circle.append(circle_info[index])
            cate_edge_points = circle_edges_points[index]
            for point in cate_edge_points:
                x,y = point
                if [x,y] not in circle_points:
                    circle_points.append([x,y])
        # 同一组点，求均值
        cate_average_x, cate_average_y, cate_average_radius, cate_sum_k = np.mean(np.array(cate_circle)[:, 0]), np.mean(np.array(cate_circle)[:, 1]), np.mean(np.array(cate_circle)[:, 2]), np.sum(np.array(cate_circle)[:,4])
        new_circle_info.append([cate_average_x, cate_average_y, cate_average_radius, len(new_circle_points), cate_sum_k])
        new_circle_points.append(circle_points)

    return new_circle_info, new_circle_points


def classifyCircle(circle_tree):
    start_template_info = []
    other_template_info = []
    code_circle_info = []
    locate_circle_info = []


    other_circle_info = []
    max_radius = 0
    for position in circle_tree:

        if len(position) == 1:
            # for circle in position[0]:
            other_circle_info.append(position)
            radius = position[0][2]
            if radius>max_radius:
                max_radius = radius
        elif len(position) == 2:
            start_template_info.append(position)
            # for radius in position:
            #     for circle in radius:
            #         start_template_info.append(circle)
            #         start_template_points.append(circle_edge_points[int(circle[3])])
        else:
            # other_template_circle = []
            # other_template_circle_p = []
            # for radius in position:
            #     for circle in radius:
            #         other_template_circle.append(circle)
            #         other_template_circle_p.append(circle_edge_points[int(circle[3])])
            other_template_info.append(position)
    for other_circle in other_circle_info:
        circle = other_circle[0]
        if abs(circle[2]-max_radius) < 5:
            locate_circle_info.append(other_circle)
        else:
            code_circle_info.append(other_circle[0])
    # return start_template_info, start_template_points, other_template_info, other_template_points, locate_circle_info, locate_circle_points, code_circle_info, code_circle_points
    return start_template_info,other_template_info, locate_circle_info,code_circle_info


def showClassifyCircle(starts, templates, locates, codes, circle_edges_points):
    classify_canvas = np.zeros((height, width, 3), np.uint8)
    for pos in starts:
        for circle in pos:
            circle_edges = circle_edges_points[int(circle[3])]
            for p in circle_edges:
                x, y = p
                classify_canvas[y, x, :] = (14, 169, 250)   #橘色
    for pos in templates:
        for circle in pos:
            circle_edges = circle_edges_points[int(circle[3])]
            for p in circle_edges:
                x, y = p
                classify_canvas[y, x, :] = (17, 17, 255)  # 红色
    for pos in codes:
        circle_edges = circle_edges_points[int(pos[3])]
        for p in circle_edges:
            x, y = p
            classify_canvas[y, x, :] = (247, 162, 17)   # 蓝色
    for pos in locates:
        for circle in pos:
            circle_edges = circle_edges_points[int(circle[3])]
            for p in circle_edges:
                x, y = p
                classify_canvas[y, x, :] = (125, 186, 94)   # 绿色
    return classify_canvas


def getCenterBySubEdges(template_center, edges_points):
    sub_centers = []
    sub_edges_points = []
    cur_edges_points = []
    nan_canvas = np.zeros((height, width, 3), np.uint8)
    for pos in template_center:
        cur_pos_centers = []
        for circle in pos:
            cur_edges = edges_points[int(circle[3])]
            cur_sub_edges = getSubEdgePoints(cur_edges)
            (s_x, s_y), (s_a, s_b), angle = cv2.fitEllipse(np.array(cur_sub_edges, dtype=np.float32))
            if np.isnan(s_x) or np.isnan(s_y) or np.isnan(s_a) or np.isnan(s_b):
                for point in cur_sub_edges:
                    ii,jj = point[0]
                    print(point[0], ii, jj)
                    i, j = int(jj), int(ii)
                    nan_canvas[i,j,:] = (0,0,255)
                # cv2.imwrite("../images/process/angle/gaussianBlur/" + imgName + '-nancanvas.png', nan_canvas)
                continue
            else:
                # print(np.float32(s_x), np.float32(s_y), np.float32(max(s_a, s_b) / 2))
                cur_pos_centers.append([np.float32(s_x), np.float32(s_y), np.float32(max(s_a, s_b) / 2)])
            sub_points = [[x,y] for [[x,y]] in cur_sub_edges]
            sub_edges_points.append(sub_points)
            cur_edges_points.append(cur_edges)
        sub_centers.append(cur_pos_centers)

    return sub_centers, sub_edges_points, cur_edges_points


def getCenterBySubEdges1(template_center, edges_points):
    sub_centers = []
    sub_edges_points = []
    cur_edges_points = []
    for pos in template_center:
        cur_pos_centers = []
        for circle in pos:
            cur_edges = edges_points[int(circle[3])]
            cur_sub_edges = getSubEdgePoints(cur_edges)
            s_x, s_y, s_r = ransac_circle_fit(cur_sub_edges, 3000, 0.8*len(cur_sub_edges), 1)
            print("ransac_res:", s_x, s_y, s_r)
                # print(np.float32(s_x), np.float32(s_y), np.float32(max(s_a, s_b) / 2))
            cur_pos_centers.append([np.float32(s_x), np.float32(s_y), np.float32(s_r)])
            # sub_points = [[x,y] for [[x,y]] in cur_sub_edges]
            # sub_edges_points.append(sub_points)
            # cur_edges_points.append(cur_edges)
        sub_centers.append(cur_pos_centers)

    return sub_centers, sub_edges_points, cur_edges_points


def showSubEdges(edges, sub_edges):
    p_x, p_y, sub_x, sub_y = [], [], [], []
    for i in range(len(edges)):
        e = edges[i]
        s = sub_edges[i]
        p_x.extend(np.array(e)[:,0])
        p_y.extend(np.array(e)[:,1])
        sub_x.extend(np.array(s)[:, 0])
        sub_y.extend(np.array(s)[:, 1])
    plt.imshow(blur, cmap="gray")
    plt.scatter(p_x, p_y, s=10, marker="*")
    plt.scatter(sub_x, sub_y, s=5, marker="*")
    plt.show()


def getGrayMoment(gray, edges):
    # 计算亚像素梯度
    dx = cv2.Sobel(gray, cv2.CV_64F, 1, 0, ksize=3)
    dy = cv2.Sobel(gray, cv2.CV_64F, 0, 1, ksize=3)

    # 定义系数矩阵
    coef_mat = np.matrix([[1, 0, 0, 0], [0.5, 0.5, 0, 0], [1 / 3, 1 / 6, 1 / 3, 0], [0.25, 1 / 3, 0.25, 1 / 4]])
    subX, subY = [], []
    # 遍历每个边缘点
    for y in range(1, gray.shape[0] - 1):
        for x in range(1, gray.shape[1] - 1):
            if edges[y, x] > 0:
                # 计算灰度矩
                m = np.zeros((4, 1))
                for p in range(4):
                    for q in range(4 - p):
                        px = x + p - 1
                        qy = y + q - 1
                        m[p + q] += coef_mat[p, q] * gray[qy, px]

                # 计算亚像素梯度
                Gx = dx[y, x]
                Gy = dy[y, x]

                # 构造函数 F
                F = lambda dx, dy: np.dot(m.transpose(),
                                          np.power(np.array([dx ** i * dy ** (3 - i) for i in range(4)]), 2))

                # 最小二乘法求解最小值点
                A = np.array([[F(1, 0), F(0, 1)], [Gx, Gy]])
                b = np.array([-F(1, 0) + m[1], -Gx])
                d = np.linalg.solve(A.astype(np.float32), b.astype(np.float32))

                # 更新边缘点坐标
                # edges[y, x] = 0
                # if abs(d[0]) <= 0.5 and abs(d[1]) <= 0.5:
                if abs(d[0]) <= 1 and abs(d[1]) <= 1:
                    subX.append(x+d[0])
                    subY.append(y+d[1])
    return subX, subY


def unSharpMask(img, sigma, amount, thresh):
    """
    图像锐化:  增强图像边缘信息，计算公式：Y = X+λZ  X为源图像，Z为校正因子，此处为高通滤波后的图像，λ为缩放因子
    :param img: 源图像
    :param radius: 高斯内核
    :param amount: λ
    :param thresh:  阈值
    :return:
    """
    lImg = cv2.GaussianBlur(img, (0,0), sigma)
    hImg = cv2.subtract(img, lImg)

    mask = hImg > thresh
    newVal = img + mask*amount*hImg/100
    newImg = np.clip(newVal, 0, 255)
    newImg = newImg.astype(np.uint8)
    return newImg





# filePath = "../documents/0625_angle_error.xlsx"
# wb = openpyxl.load_workbook(filePath)
# sheet = wb['Sheet1']
# img_name_list = [
#     "marker_0-9", "marker_0-20-9", "marker_1-8", "marker_1-21-10", "marker_2-9", "marker_2-22-10",
#     "marker_3-9", "marker_23-10", "marker_4-9", "marker_24-10", "marker_5-9", "marker_25-9", "marker_6-10", "marker_26-9",
#     "marker_7-8", "marker_17-9", "marker_8-0", "marker_18-10", "marker_9-7", "marker_19-9", "marker_10-8", "marker_20-10",
#     "marker_11-9", "marker_21-8", "marker_12-9", "marker_22-8"
# ]
# r = 1
# filePath = "../documents/0627_location.xlsx"
# wb = openpyxl.load_workbook(filePath)
# sheet = wb['Sheet1']
# r=43
# img_name_list = ["marker_A-0", "marker_B-2", "marker_C-2", "marker_D-5", "marker_O"]
# for imgName in img_name_list:
#     print(imgName)
#     r += 1
imgName = "marker_20-clahe-usm"
# img = cv2.imread('../images/process/angle/imgAngle/'+imgName+'.bmp', 0)
# img = cv2.imread('../images/process/angle/imgLocation/'+imgName+'.bmp', 0)
img = cv2.imread('../images/process/0428/'+imgName+'.png', 0)
height, width = img.shape[0], img.shape[1]
# blur = cv2.medianBlur(img, 3)
blur = cv2.GaussianBlur(img, (0, 0), 2.2)
# clahe = cv2.createCLAHE(2, (16,16))
# blur = clahe.apply(blurred)
# cv2.imwrite("../images/process/angle/medianBlur/" + imgName + '-blur.png', blur1)
# cv2.imwrite("../images/process/angle/gaussianBlur/" + imgName + '-blur.png', blur)
# img = cv2.imread("../images/process/0428/" + imgName + ".png", 0)
# height, width = img.shape[0], img.shape[1]
# blur = img.copy()
# cv2.imwrite("../images/process/angle/gaussianBlur/" + imgName + '-blur.png', blur)
sobelx = cv2.Sobel(blur, cv2.CV_64F, 1, 0)
sobely= cv2.Sobel(blur, cv2.CV_64F, 0, 1)

mag, angle = cv2.cartToPolar(sobelx, sobely, angleInDegrees=True)

grad_mag = cv2.normalize(mag, None, 0, 255, cv2.NORM_MINMAX, cv2.CV_8UC1)
# plt.hist(grad_mag.ravel(), 256)
# plt.show()
th, _ = cv2.threshold(grad_mag, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
# print("th:", th)
tl = max(0, int(th * 0.3))
print(th, tl)
edges = cv2.Canny(blur, tl,th-10)
cv2.imwrite("../images/process/summary/gaussianOstuCanny/"+imgName+"_edges.png", edges)
_, sobel_thresh = cv2.threshold(grad_mag, 10,255, cv2.THRESH_BINARY)
cv2.imwrite("../images/process/summary/sobelThresh/"+imgName+"_sobel_edges.png", sobel_thresh)
sobel_angle = np.arctan(sobely, sobelx)

rows, cols = np.where(edges==255)
edges_points = [[x,y] for x,y in list(zip(cols, rows))]
edge_canvas = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
for pp in edges_points:
    x, y = pp
    edge_canvas[y,x,:] = (0,0,255)
cv2.imwrite("../images/process/summary/gaussianOstuCanny/"+imgName+"_edges2.png", edge_canvas)
print("边缘点的个数为：", len(edges_points))
# 去除分叉的边缘（可消除圆弧交叉的情况）
no_branch = delBranch(edges, 60)
cv2.imwrite("../images/process/summary/delBranch/" + imgName + '-no_branch.png', no_branch)

# 圆弧检测
filter_cnts = np.zeros(edges.shape, dtype=np.uint8)
contours, _ = cv2.findContours(no_branch, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
circle_info, circle_edge_points = extractCircles(contours, no_branch, edges_points)
# print(len(circle_edge_points))
## 圆弧提取可视化
showExtractArc(circle_edge_points, "extractArc")
## 圆弧分组-----将属于同一个圆的圆弧分到同一组
group_circles_info, group_circles_points = groupCircles(circle_info, circle_edge_points, 1.5, 1.5)
# group_circle_info中的元素为属于同一个圆的[圆心坐标平均值，半径平均值，对应group_circle_points中的边缘索引， k之和]
showExtractArc(group_circles_points, "groupArc")



# 圆弧分类-----按位置分
category_index = classifyArcByPosition(group_circles_info)
showClassifyArcByPosition(category_index, group_circles_info, group_circles_points)
# category_index列表，与group_circle_info长度相等，每一个元素代表group_circle_info相同索引的元素所属的类别
all_circle_tree = classifyArcByRadius(group_circles_info, category_index)
showClassifyArcByRadius(all_circle_tree, group_circles_points, "classifyByRadius")

simplify_circle_tree = filterArcByMaxK(all_circle_tree)
showFilterArc(simplify_circle_tree, group_circles_points, "simplifyArc")



start_template_info, other_template_info,locate_circle_info, code_circle_info = classifyCircle(simplify_circle_tree)
res_canvas = showClassifyCircle(start_template_info, other_template_info, locate_circle_info, code_circle_info, group_circles_points)
cv2.imwrite("../images/process/summary/templateCircle/" + imgName + '-category.png', res_canvas)


# 填补编码圆
int_thresh = fillIntCircle(no_branch, code_circle_info)

# time1 = time.time()
# start_template_centers, start_template_sub_edges, start_template_edges = getCenterBySubEdges(start_template_info, group_circles_points)
# other_template_centers, other_template_sub_edges, other_template_edges = getCenterBySubEdges(other_template_info, group_circles_points)
# print("locate_circle_info:", locate_circle_info)
# locate_centers, locate_sub_edges, locate_edges = getCenterBySubEdges(locate_circle_info, group_circles_points)
# time2 = time.time()
# print("耗时：", time2-time1)

# print(start_template_sub_edges)
# showSubEdges(start_template_edges, start_template_sub_edges)
# showSubEdges(other_template_edges, other_template_sub_edges)
# showSubEdges(locate_edges, locate_sub_edges)

# 使用均值求出模板点、定位圆的中心坐标点---粗定位
start_circle_centers = calCentersByAverage(start_template_info)
template_circle_centers = calCentersByAverage(other_template_info)
locate_circle_centers = calCentersByAverage(locate_circle_info)
# start_circle_centers = calCentersByAverage(start_template_centers)
# template_circle_centers = calCentersByAverage(other_template_centers)
# locate_circle_centers = calCentersByAverage(locate_centers)
# print(template_circle_centers)
# print("locate_circle_centers", locate_circle_centers)
# 识别N1, N2, N3
## 设计坐标
# 根据设计图的特征，假设模板点坐标如下：
src_n0 = [50, 50]
src_n1 = [250, 50]
src_n2 = [50, 250]
src_n3 = [250, 250]
src_lp = [150, 150]
source_points = np.array([src_n0, src_n1, src_n2, src_n3], dtype=np.float32)
# 整数编码圆中心的设计坐标
code_points = generateCodePoints(src_n0, 200)  # 根据设计图的比例生成编码点的中心坐标列表
# band_angles = calculateBandAnglesByTemplatePoint(src_lp, src_n0, src_n1, src_n2, src_n3)    # 计算设计图上每个编码带的起始角度、终止角度
# 小数编码块中心的设计坐标
band_angles = calculateBandAngleByAverage(10)
bandCenters = getBandCenter(band_angles, src_lp, 200)  # 计算每个编码带中间位置的设计坐标
# 从source_points中任选3点的组合方式
source_combinations = np.array(list(itertools.combinations(source_points, 3)))


drawCircle = cv2.cvtColor(filter_cnts, cv2.COLOR_GRAY2BGR)

for n0 in start_circle_centers:
    cv2.putText(drawCircle, 'n0', (int(n0[0] + 50), int(n0[1] - 20)), cv2.FONT_HERSHEY_SIMPLEX, 1, (247, 162, 17), 2)
    cv2.circle(drawCircle, (int(n0[0]), int(n0[1])), 2, (247, 162, 17), -1)
    cv2.imwrite("../images/process/summary/templateCircle/" + imgName + '-recognize.png', drawCircle)
    if len(template_circle_centers)<2 or len(locate_circle_centers)<1:
        continue
    lp, n1, n2, n3 = matchPoint(template_circle_centers, n0, locate_circle_centers)

    ## 可视化
    cv2.putText(drawCircle, 'n1', (int(n1[0]+50), int(n1[1]-20)), cv2.FONT_HERSHEY_SIMPLEX, 1, (125, 186, 94), 2)
    cv2.circle(drawCircle, (int(n1[0]), int(n1[1])), 2, (125, 186, 94), -1)
    cv2.putText(drawCircle, 'n2', (int(n2[0]+50), int(n2[1]-20)), cv2.FONT_HERSHEY_SIMPLEX, 1, (17,17, 255), 2)
    cv2.circle(drawCircle, (int(n2[0]), int(n2[1])), 2, (17,17, 255), -1)
    cv2.putText(drawCircle, 'n3', (int(n3[0]+50), int(n3[1]-20)), cv2.FONT_HERSHEY_SIMPLEX, 1, (14, 169, 250), 2)
    cv2.circle(drawCircle, (int(n3[0]), int(n3[1])), 2, (14, 169, 250), -1)
    cv2.imwrite("../images/process/summary/templateCircle/" + imgName + '-recognize.png', drawCircle)

    # 从n0, n1, n2, n3中找到与combinations组合中与src_x对应的点
    dist_points = np.array([n0, n1, n2, n3], dtype=np.float32)
    dist_combinations = getDistCombinations(source_combinations, source_points, dist_points)

    angle = 0.0
    transform_matrix = np.zeros((2,3), np.float32)
    max_dist = 10.0

    for i in range(len(source_combinations)):
        MM = calculateAffineMatrix(source_combinations[i], dist_combinations[i])
        s_remain_p, d_remain_p = getRemainPoint(source_combinations[i], source_points, dist_points)
        cal_dist_x = round(MM[0][0] * s_remain_p[0] + MM[0][1] * s_remain_p[1] + MM[0][2], 4)
        cal_dist_y = round(MM[1][0] * s_remain_p[0] + MM[1][1] * s_remain_p[1] + MM[1][2], 4)
        dist_err = np.sqrt((cal_dist_x-d_remain_p[0])**2 + (cal_dist_y-d_remain_p[1])**2)
        print(dist_err, 135 - np.arctan2(MM[1, 0], MM[0, 0]) * 180 / np.pi)
        if dist_err<max_dist:
            max_dist = dist_err
            angle = 135 - np.arctan2(MM[1, 0], MM[0, 0]) * 180 / np.pi
            print("min:", dist_err, angle)
            transform_matrix = MM
    angle = angle if angle>0 else 360+angle
    print("标识的角度为：", angle)
    ## 整数部分解码
    side_length = np.sqrt((n0[0]-n1[0])**2+(n0[1]-n1[1])**2)
    int_points = cvtCodePoints1(code_points, transform_matrix)
    int_str = getCodeVal(int_thresh, int_points)
    ## 小数部分解码
    dec_points = cvtCodePoints1(bandCenters, transform_matrix)
    # 提取出所有的圆环带轮廓
    circular_thresh = extractCircularEdges(sobel_thresh, side_length)
    dec_value = getCodeVal(circular_thresh, dec_points)
    # dec_value1 = decodedByGray(dec_points, lp, side_length, blur)
    # print("小数:", int(dec_value, 2), int(dec_value1, 2), dec_value, dec_value1)
    print('标记点的解码结果为:{}.{}'.format(int(int_str, 2), str(int(dec_value, 2)).zfill(3)))


    # 计算中心坐标
    print("定位圆中心：", lp)
    [x_o2, y_o2] = calculateCrossPoint([n0, n3], [n1, n2])
    average_x = (x_o2 + lp[0]) / 2
    average_y = (y_o2 + lp[1]) / 2
    print('标记点的中心坐标为:{}'.format([average_x, average_y]))

    ## 计算对准相机中心平台需要移动的距离
    # move_x = round(97*0.545/83 * (average_y - 1057), 4)
    # move_y = round(97*0.545/83 * (average_x - 960), 4)
    # move_x = round(95*0.53/83 * (average_y - 1071), 4)
    # move_y = round(95*0.53/83 * (average_x - 1214), 4)
    move_x = round(95 * 0.54 / 82 * (lp[1]-1068.5)*0.001, 6)
    move_y = round(95 * 0.54 / 82 * (lp[0]-1214)*0.001, 6)
    print("平台X方向移动{}μm,Y方向移动{}μm".format(move_x, move_y))
    L = int(int_str, 2) + int(dec_value, 2)*0.001
    next_x = -L * math.sin(math.radians(angle))
    next_y = L * math.cos(math.radians(angle))
    print("找到下一个标记点：平台X方向移动{}μm，Y方向移动{}μm".format(next_x, next_y))
    # sheet.cell(row=r, column=10, value=lp[0])
    # sheet.cell(row=r, column=11, value=lp[1])
    # sheet.cell(row=r, column=6, value=move_x)
    # sheet.cell(row=r, column=7, value=move_y)
    # sheet.cell(row=r, column=6, value=move_x)
    # sheet.cell(row=r, column=7, value=move_y)
    # original_x = sheet.cell(row=r, column=2).value
    # original_y = sheet.cell(row=r, column=3).value
    # real_x = original_x + move_x
    # real_y = original_y + move_y
    # sheet.cell(row=r, column=10, value=real_x)
    # sheet.cell(row=r, column=11, value=real_y)
    # sheet.cell(row=r, column=17, value=135 -angle)
    # sheet.cell(row=r, column=18, value=135 -angle2)
    # sheet.cell(row=r, column=19, value=135 -angle1)


# wb.save(filePath)


