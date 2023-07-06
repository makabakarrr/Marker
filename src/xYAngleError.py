import openpyxl
import numpy as np
import matplotlib.pyplot as plt


filePath = '../documents/mass/0704_center.xlsx'
wb = openpyxl.load_workbook(filePath)

sheet = wb['Sheet1']


p_err_x, p_err_y, p_err_angle = [], [], []
sub_err_x, sub_err_y, sub_err_angle = [], [], []

for r in range(2, 32):
    normal_x = sheet.cell(row=r, column=2).value
    normal_y = sheet.cell(row=r, column=3).value
    normal_angle = sheet.cell(row=r, column=4).value
    p_x = sheet.cell(row=r, column=5).value
    p_y = sheet.cell(row=r, column=6).value
    p_angle = sheet.cell(row=r, column=7).value
    sub_x = sheet.cell(row=r, column=8).value
    sub_y = sheet.cell(row=r, column=9).value
    sub_angle = sheet.cell(row=r, column=10).value
    k = sheet.cell(row=r, column=11)
    l = sheet.cell(row=r, column=12)
    m = sheet.cell(row=r, column=13)
    n = sheet.cell(row=r, column=14)
    o = sheet.cell(row=r, column=15)
    p = sheet.cell(row=r, column=16)
    # p_x_err = abs(normal_x-p_x)
    # p_y_err = abs(normal_y-p_y)
    # sub_x_err = abs(normal_x-sub_x)
    # sub_y_err = abs(normal_y-sub_y)
    # p_angle_err = abs(normal_angle-p_angle)
    # sub_angle_err = abs(normal_angle-sub_angle)
    k.value = abs(normal_x-p_x)
    l.value = abs(normal_y-p_y)
    n.value = abs(normal_x-sub_x)
    o.value = abs(normal_y-sub_y)
    m.value = abs(normal_angle-p_angle)
    p.value = abs(normal_angle-sub_angle)
    p_err_x.append(k.value)
    p_err_y.append(l.value)
    p_err_angle.append(m.value)
    sub_err_x.append(n.value)
    sub_err_y.append(o.value)
    sub_err_angle.append(p.value)

wb.save(filePath)

p_err_x_mean, p_err_y_mean, p_err_angle_mean = np.mean(p_err_x), np.mean(p_err_y), np.mean(p_err_angle)
sub_err_x_mean, sub_err_y_mean, sub_err_angle_mean = np.mean(sub_err_x), np.mean(sub_err_y), np.mean(sub_err_angle)
plt.subplot(131)
plt.plot(range(len(p_err_x)), p_err_x, color="#4A235A", marker="o",markeredgecolor="white", linestyle="-", label="pixel")
plt.axhline(p_err_x_mean,color="#4A235A", linestyle="--")
plt.plot(range(len(sub_err_x)), sub_err_x, color="#D7263D", marker="s",markeredgecolor="white", linestyle="-", label="sub_pixel")
plt.axhline(sub_err_x_mean, color="#D7263D", linestyle="--")
plt.ylabel('X-error')
plt.xlabel('Marker')
plt.legend()
plt.subplot(132)
plt.plot(range(len(p_err_y)), p_err_y, color="#4A235A", marker="o",markeredgecolor="white", linestyle="-", label="pixel")
plt.axhline(p_err_y_mean,color="#4A235A", linestyle="--")
plt.plot(range(len(sub_err_y)), sub_err_y, color="#D7263D", marker="s",markeredgecolor="white", linestyle="-", label="sub_pixel")
plt.axhline(sub_err_y_mean, color="#D7263D", linestyle="--")
plt.ylabel('Y-error')
plt.xlabel('Marker')
plt.legend()
plt.subplot(133)
plt.plot(range(len(p_err_angle)), p_err_angle, color="#4A235A", marker="o",markeredgecolor="white", linestyle="-", label="pixel")
plt.axhline(p_err_angle_mean,color="#4A235A", linestyle="--")
plt.plot(range(len(sub_err_angle)), sub_err_angle, color="#D7263D", marker="s",markeredgecolor="white", linestyle="-", label="sub_pixel")
plt.axhline(sub_err_angle_mean, color="#D7263D", linestyle="--")
plt.ylabel('Angle-error')
plt.xlabel('Marker')
plt.legend()

plt.show()
