import random
import numpy as np
import matplotlib.pyplot as plt
import openpyxl

plt.rcParams['font.sans-serif']=['kaiti']

m = 1
l_min = 350
l_max = 850
l = l_min
l_list = [350, 360, 370, 380, 390, 400]   #
R = 96


def six_variable_function(xx1, xx2, xx3, xx4, xx5, xx6):
    y = 0.0
    x1, x2, x3, x4, x5, x6 = np.float(xx1), np.float(xx2), np.float(xx3), np.float(xx4), np.float(xx5), np.float(xx6)
    x3, x4 = np.radians(x3), np.radians(x4)
    for l in range(350, 851):
        l = round(l*0.001, 4)
        temp = np.clip(m*l/x5 - np.sin(x3), -1, 1)
        y1 = np.arcsin(temp)  # 角度的单位为弧度---β(λq)
        y2 = x2 / (np.cos(x4-y1))
        y3 = np.cos(x3)**2 / x1 + np.cos(y1)**2 / y2 - (np.cos(x3)+np.cos(y1)) / R
        y4 = (y3 + m*l/x5 * x6)**2
        # print("y1", l, x5, m*l/x5, np.sin(x3), m*l/x5 - np.sin(x3))
        # print("y1y2y3y4", y1, y2, y3, y4)
        y += y4

    return y



# 定义遗传算法参数
population_size = 70   # 种群数量
max_generations = 50    # 迭代代数
mutation_rate = 0.5     # 变异率
elite_size = 10          # 精英数量


# 定义变量范围和编码长度
variable_ranges = [(0, 200), (0, 200), (-90, 90), (-90, 90), (0.5, 5), (-10, 10)]
variable_length = [20, 20, 15, 15, 16, 18]



# 初始化种群
# 生成初始种群
def generate_population(size):
    population = []
    for _ in range(population_size):
        individual = []
        for i in range(6):
            individual.append(''.join(random.choice('01') for _ in range(variable_length[i])))
        population.append(individual)
    return population


def decoded_x(individual):
    # x1 = variable_ranges[0][0] + int(individual[0], 2) * (variable_ranges[0][1] - variable_ranges[0][0]) / (
    #             np.power(2, variable_length[0]) - 1)
    # x2 = variable_ranges[1][0] + int(individual[1], 2) * (variable_ranges[1][1] - variable_ranges[1][0]) / (
    #             np.power(2, variable_length[1]) - 1)
    # x3 = variable_ranges[2][0] + int(individual[2], 2) * (variable_ranges[2][1] - variable_ranges[2][0]) / (
    #             np.power(2, variable_length[2]) - 1)
    # x4 = variable_ranges[3][0] + int(individual[3], 2) * (variable_ranges[3][1] - variable_ranges[3][0]) / (
    #             np.power(2, variable_length[3]) - 1)
    # x5 = variable_ranges[4][0] + int(individual[4], 2) * (variable_ranges[4][1] - variable_ranges[4][0]) / (
    #             np.power(2, variable_length[4]) - 1)
    # x6 = variable_ranges[5][0] + int(individual[5], 2) * (variable_ranges[5][1] - variable_ranges[5][0]) / (
    #             np.power(2, variable_length[5]) - 1)

    x1, x2, x3, x4, x5, x6 = [variable_ranges[i][0] + int(individual[i], 2) * (variable_ranges[i][1] - variable_ranges[i][0]) / (
                np.power(2, variable_length[i]) - 1) for i in range(6)]

    return x1, x2, x3, x4, x5, x6


# 定义适应度函数 (I20越小，适应度越大)
def fitness_function(individual):
    x1, x2, x3, x4, x5, x6 = decoded_x(individual)
    # print("变量取值：", x1, x2, x3, x4, x5, x6)
    print("函数值：", six_variable_function(x1, x2, x3, x4, x5, x6))

    return  six_variable_function(x1, x2, x3, x4, x5, x6)


# 选择操作：根据适应度排序选择部分个体（适应度最大）
def selection(population):
    sorted_population = sorted(population, key=fitness_function)
    return sorted_population[:elite_size]


# 交叉操作：使用两点交叉生成新个体
def crossover(parent1, parent2):
    point1 = random.randint(0, 5)
    point2 = random.randint(0, 5)
    if point1 > point2:
        point1, point2 = point2, point1
    child = parent1[:point1] + parent2[point1:point2] + parent1[point2:]
    return child


# 变异操作：对个体的染色体进行变异
def mutate(chromosome):
    for i in range(6):
        if random.random() < mutation_rate: # 根据变异率进行变异
            chromosome_list = list(chromosome[i])
            mutation_point = random.randint(0, variable_length[i] - 1)
            chromosome_list[mutation_point] = '1' if chromosome_list[mutation_point] == '0' else '0'
            chromosome[i] = ''.join(chromosome_list)
    return chromosome


# 遗传算法主流程
def genetic_algorithm():
    best_list = []
    population = generate_population(population_size)

    for j in range(max_generations):
        new_population = selection(population)
        if j!=0:
            best_x = new_population[0]
            best_fitness = fitness_function(best_x)
            best_list.append(best_fitness)

        while len(new_population) < population_size:
            parent1 = random.choice(population)
            parent2 = random.choice(population)
            child = crossover(parent1, parent2) # 随机选取两个个体进行交叉
            child = mutate(child)   # 变异
            new_population.append(child)

        population = new_population


    best_individual = selection(population)[0]
    return best_individual, best_list


# 运行遗传算法
best_solution, fitness_list = genetic_algorithm()
x1, x2, x3, x4, x5, x6 = decoded_x(best_solution)
best_fitness = six_variable_function(x1, x2, x3, x4, x5, x6)

# 输出结果
print("Best solution:", (x1, x2, x3, x4, x5, x6))
print("Best fitness:", best_fitness)
filePath = '../documents/genetic.xlsx'
wb = openpyxl.load_workbook(filePath)

sheet = wb['Sheet1']
row = sheet.max_row + 1
sheet.cell(row=row, column=1, value=row-1)
sheet.cell(row=row, column=2, value=round(x1, 4))
sheet.cell(row=row, column=3, value=round(x2, 4))
sheet.cell(row=row, column=4, value=round(x3, 4))
sheet.cell(row=row, column=5, value=round(x4, 4))
sheet.cell(row=row, column=6, value=round(x5, 4))
sheet.cell(row=row, column=7, value=round(x6, 4))
wb.save(filePath)

plt.plot(fitness_list)
plt.title("适应度曲线")
plt.savefig("../images/genetic/"+row+".png")
plt.show()





