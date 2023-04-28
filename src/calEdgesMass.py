import cv2
import numpy as np
import math
import openpyxl

from customFunction import *
from utils import calculateCrossPoint, getAngle

from skimage import morphology
from collections import deque

# 根据梯度方向获取下一个边缘点
def get_next_edge_point(point, angle):
    k1 = np.array([np.cos(np.radians(angle + 90)), np.sin(np.radians(angle + 90))])
    k2 = np.array([np.cos(np.radians(angle - 90)), np.sin(np.radians(angle - 90))])
    next_point1 = np.round(np.array(point) + k1).astype(int)
    next_point2 = np.round(np.array(point) + k2).astype(int)
    return [next_point1, next_point2]

def get_next_edge_point1(point, gradient_angle):
    x, y = point
    basic_direction = round(gradient_angle / 45.0) % 4
    if basic_direction == 0:
        return (x, y + 1) if refine[x, y + 1] < 255 else (x, y - 1)
    elif basic_direction == 1:
        return (x - 1, y - 1) if refine[x - 1, y - 1] < 255 else (x + 1, y + 1)
    elif basic_direction == 2:
        return (x + 1, y) if refine[x + 1, y] < 255 else (x - 1, y)
    elif basic_direction == 3:
        return (x + 1, y - 1) if refine[x + 1, y - 1] < 255 else (x + 1, y - 1)



def detectEndPoints(cnt):
    que = deque()
    for point in cnt:
        x, y = point[0]
        neighborhood = refine[y - 1:y + 2, x - 1:x + 2]
        if np.sum(neighborhood) == 510:  # 3*3窗口内只有一个相邻像素为非零像素
            que.append([x, y])
    return que


def connectEdges(que, image):
    while len(que) > 0:
        point = que.popleft()  # 取出端点
        center_x, center_y = point
        cv2.circle(mask_canvas, (center_x, center_y), 1, (0,0,255),-1)
        next_points = get_next_edge_point(point, grad_angle[point[1]][point[0]])
        x1, y1 = next_points[1][0], next_points[1][1]
        next_point = next_points[0] if image[y1][x1] > 0 else next_points[1]
        x, y = next_point

        if not np.array_equal(point, next_point) and sobel_thresh[y][x] == 255 and image[y][x] == 0:
            image[y][x] = 255
            que.append(next_point)

def region_growing(img, seed):
    """
    :param img: 要处理的单通道图像
    :param seed: 包含种子点的二元组 (x,y) 坐标。从这个点开始向外扩散。
    :return: 区域生长后的掩码图像
    """

    # 创建空掩码图像，初始化为全零
    h, w = img.shape[:2]
    mask = np.zeros((h+2, w+2), np.uint8)

    # 参数设定
    connectivity = 4  # 连通性
    flags = connectivity
    lo_diff = 3  # 差异阈值loDiff
    up_diff = 3  # 差异阈值upDiff

    # 生长区域
    cv2.floodFill(img, mask, seed, newVal=255, loDiff=lo_diff, upDiff=up_diff, flags=flags)

    # 返回掩码图像
    return mask[1:-1, 1:-1]


def unsharpMask(img, sigma, amount, thresh):
    """
    图像锐化:  增强图像边缘信息，计算公式：Y = X+λZ  X为源图像，Z为校正因子，此处为高通滤波后的图像，λ为缩放因子
    :param img: 源图像
    :param radius: 高斯内核
    :param amount: λ
    :param thresh:  阈值
    :return:
    """
    lImg = cv2.GaussianBlur(img, (0,0), sigma)
    hImg = cv2.subtract(img, lImg)

    mask = hImg > thresh
    newVal = img + mask*amount*hImg/100
    newImg = np.clip(newVal, 0, 255)
    newImg = newImg.astype(np.uint8)
    # h, w = img.shape
    # for i in range(0, h):
    #     for j in range(0, w):
    #         val = hImg[i][j]
    #         if  val> thresh:
    #             newVal = img[i][j] + amount*val/100
    #             img[i][j] = 0 if newVal < 0 else (255 if newVal > 255 else newVal)
    return newImg



filePath = '../documents/marker_center_0428.xlsx'
# # 打开 XLSX 文件
wb = openpyxl.load_workbook(filePath)
#
# # 选择工作表
sheet = wb['Sheet1']


for m in range(0, 190):
    imgName = "marker_" + str(m)
    img = cv2.imread('../images/process/0428/' + imgName + '.png', 0)

    ## 图像增强
    clahe = cv2.createCLAHE(clipLimit=1.5, tileGridSize=(32, 32))
    enhance = clahe.apply(img)
    # cv2.imwrite("../images/process/0426/process/" + imgName + "-enhance.bmp", enhance)

    blurred = cv2.GaussianBlur(enhance, (0, 0), 10)
    # usm = cv2.addWeighted(enhance, 1.8, blurred, -0.3, 0)
    usm = unsharpMask(enhance, 10, 60, 0)
    # cv2.imwrite("../images/process/0426/process/" + imgName + "-usm.bmp", usm)
    blurred1 = cv2.GaussianBlur(usm, (0, 0), 1)

    sobelx = cv2.Sobel(blurred1, cv2.CV_64F, 1, 0, ksize=3)
    sobely = cv2.Sobel(blurred1, cv2.CV_64F, 0, 1, ksize=3)
    # # 计算梯度幅值和方向
    grad_mag, grad_angle = cv2.cartToPolar(sobelx, sobely, angleInDegrees=True)
    grad_mag = cv2.normalize(grad_mag, None, 0, 255, cv2.NORM_MINMAX, cv2.CV_8UC1)
    _, sobel_thresh = cv2.threshold(grad_mag, 30, 255, cv2.THRESH_BINARY)
    # cv2.imwrite("../images/process/0426/process/" + imgName + '-grad_thresh.bmp', sobel_thresh)

    edges = cv2.Canny(blurred1, 50, 150)
    # cv2.imwrite("../images/process/0426/process/" + imgName + "-edges.bmp", edges)

    edges[edges == 255] = 1
    skeleton0 = morphology.skeletonize(edges)
    refine = skeleton0.astype(np.uint8) * 255
    # cv2.imwrite("../images/process/0426/process/" + imgName + "-refine.bmp", refine)

    circleCnts, _ = cv2.findContours(refine, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)  # cv2.RETR_TREE
    circularCnts, _ = cv2.findContours(sobel_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)  # cv2.RETR_TREE

    # 画布
    mask1 = np.zeros(edges.shape[:2], dtype=np.uint8)
    mask2 = np.zeros(edges.shape[:2], dtype=np.uint8)
    mask_canvas = cv2.cvtColor(refine, cv2.COLOR_GRAY2BGR)
    mask_canvas1 = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
    mask_canvas2 = cv2.cvtColor(edges, cv2.COLOR_GRAY2BGR)

    for index in range(len(circleCnts)):
        cnt = circleCnts[index]
        length = cv2.arcLength(cnt, False)
        if length > 50:  # 过滤掉长度不够的轮廓
            area = cv2.contourArea(cnt, False)
            (x, y), (a, b), angle = cv2.fitEllipse(cnt)
            if (area < 10 and length > 1.6 * math.pi * a) or (
                    area > 10 and abs(4 * math.pi * area / ((math.pi * a) ** 2) - 1) < 0.1):  # 筛选圆形边缘（包含闭合的和断开的）
                cv2.drawContours(mask1, [cnt], -1, (255, 255, 255), 1)
                que = detectEndPoints(cnt)  # 检测端点
                if len(que) >= 2:
                    connectEdges(que, mask1)
    # cv2.imwrite("../images/process/0426/process/" + imgName + "-circle.bmp", mask1)

    contours_mask, hierarchy_mask = cv2.findContours(mask1, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)  # cv2.RETR_TREE
    # # # 将轮廓按照层级进行分类
    contour_dict = {}
    for i, c in enumerate(contours_mask):
        level = 0
        next_index = hierarchy_mask[0][i][3]
        while next_index != -1:
            level += 1
            next_index = hierarchy_mask[0][next_index][3]

        if level not in contour_dict:
            contour_dict[level] = [c]
        else:
            contour_dict[level].append(c)
    #
    colors_list = [(255, 255, 255), (0, 0, 0)]
    for k, v in contour_dict.items():
        # 一条闭合的曲线有两个轮廓
        if k % 2:
            cv2.drawContours(mask1, v, -1, colors_list[((k // 2) % 2)], -1)
    # cv2.imwrite("./images/process/0411-pm/cal-edges/" + imgName + "-mask1-fill.bmp", mask1)

    for i in range(0, len(circularCnts)):
        c = circularCnts[i]
        area = cv2.contourArea(c, False)
        length = cv2.arcLength(c, False)
        if area > 400:
            cv2.drawContours(mask_canvas2, [c], -1, (0, 255, 0), 1)
            (x, y), (a, b), angle = cv2.fitEllipse(c)
            k = 4 * math.pi * area / ((math.pi * a) ** 2)
            # print(i, area, length, a, b, abs(a - b), abs(k - 1))
            cv2.putText(mask_canvas2, str(i), (int(x), int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 0, 0), 1)
            cv2.ellipse(mask_canvas2, (int(x), int(y)), (int(a / 2), int(b / 2)), angle, 0, 360, (0, 0, 255), 1)

            # if (abs(a-b)>10 and max(a,b)<900) or min(a,b)>100:    # 筛选圆环带边缘
            if max(a, b) < 500 and min(a, b) > 20 and (abs(k - 1) > 0.2) or 10<abs(a-b)<100:
                cv2.drawContours(mask2, [c], -1, (255, 255, 255), -1)

    # cv2.imwrite("../images/process/0426/process/" + imgName + "-circular.bmp", mask2)
    # cv2.imwrite("../images/process/0426/process/" + imgName + "-ellipse.bmp", mask_canvas2)

    ret, thresh = cv2.threshold(mask1 + mask2, 0, 255, cv2.THRESH_OTSU)

    ## 轮廓检测：
    contours, hierarchy = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)

    ## 筛选圆形轮廓
    # circle_cnts = getCircleCntsByArea(contours, 300)    # 通过外接圆面积筛选
    circle_cnts = getCircleCntsByRoundness(contours)  # 通过圆度筛选
    drawCircle = cv2.cvtColor(thresh, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(drawCircle, circle_cnts, -1, (0, 0, 255), 2)
    # showPic("circle", drawCircle)

    # 对圆形轮廓进行分类
    cate1, center1, cate2, center2, cate3, center3 = cateCircles(circle_cnts, thresh)

    if len(center2) < 1 or len(center3) < 1 or len(center1) < 3:
        print("没有检测到标记点！")
        continue

    # 从非模板点中将定位圆的信息分离出来,分离后cate3、center3内值包含编码圆和方向圆的信息
    locate_circle, locate_center = getLocateCircles(cate3, center3)
    if len(locate_center) < 1:
        print("没有检测到定位圆")
        continue

    # 根据设计图的特征，假设模板点坐标如下：
    src_n0 = [50, 50]
    src_n1 = [250, 50]
    src_n2 = [50, 250]
    src_n3 = [250, 250]
    src_lp = [150, 150]
    source_points = np.array([src_n0, src_n1, src_n2, src_n3], np.float32)
    code_points = generateCodePoints(src_n0, 200)  # 根据设计图的比例生成编码点的中心坐标列表
    # band_angles = calculateBandAnglesByTemplatePoint(src_lp, src_n0, src_n1, src_n2, src_n3)    # 计算设计图上每个编码带的起始角度、终止角度
    band_angles = calculateBandAngleByAverage(10)
    bandCenters = getBandCenter(band_angles, src_lp, 200)  # 计算每个编码带中间位置的设计坐标

    for i in range(0, len(center2)):  # 有几个N0说明有几个标记点
        N0 = center2[i]
        lp, n1, n2, n3 = matchPoint(center1, N0, locate_center)  # 查找与N0属于同一个标记点的模板点与定位圆
        cv2.putText(drawCircle, 'n1', (int(n1[0]), int(n1[1])), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
        cv2.putText(drawCircle, 'n2', (int(n2[0]), int(n2[1])), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
        dst_points = np.array([N0, n1, n2, n3], np.float32)  # 四个模板点的图像坐标
        # 根据变换矩阵求取图像坐标再进行解码
        M = calculatePerspectiveMatrix(source_points, dst_points)  # 计算投影变换矩阵  设计坐标→图像坐标
        ## 1 整数部分的解码
        # a 根据变换矩阵和编码点的设计坐标计算编码点的图像坐标
        transform_code_points = cvtCodePoints(code_points, M)
        # b 在预处理后的图像中进行解码
        int_value = getCodeVal(thresh, transform_code_points)
        ## 2 小数部分的解码
        # a 坐标转换
        transform_band_points = cvtCodePoints(bandCenters, M)
        # b 解码
        dec_value = getCodeVal(thresh, transform_band_points)

        decode = int(int_value, 2)+int(dec_value, 2)*0.001
        sheet.cell(row=m + 2, column=7, value=decode)

        # print('第{}个标记点的解码结果为:{}.{}'.format(i + 1, int(int_value, 2), str(int(dec_value, 2)).zfill(3)))
        [x_o2, y_o2] = calculateCrossPoint([N0, n3], [n1, n2])
        average_x = (x_o2 + lp[0]) / 2
        average_y = (y_o2 + lp[1]) / 2
        # print('第{}个标记点的中心坐标为:{}'.format(i + 1, [average_x, average_y]))
        sheet.cell(row=m + 2, column=4, value=round(average_x, 4))
        sheet.cell(row=m + 2, column=5, value=round(average_x, 4))

        ## 方向识别
        # 取n1-n3的中点
        angle = getAngle(lp, N0)
        # print("方向为：", angle)
        sheet.cell(row=m + 2, column=6, value=round(angle, 4))

        ## 计算对准相机中心平台需要移动的距离
        # move_x = 81 / 51.36 * (average_y - 1071)
        # move_y = 81 / 51.36 * (average_x - 1210)
        # print("平台X方向移动{}μm,Y方向移动{}μm".format(move_x, move_y))

wb.save(filePath)






# cv2.imwrite("../images/process/marker/process/"+imgName+"-contours.bmp", mask_canvas1)
#
# # d_knernal = cv2.getStructuringElement(cv2.MORPH_RECT, (3,3))
# # e_knernal = cv2.getStructuringElement(cv2.MORPH_CROSS, (2,2))
# # cv2.dilate(mask1, d_knernal, mask1)
# # cv2.erode(mask1, e_knernal, mask1)
# # cv2.imwrite("./images/0411-pm/cal-edges/"+imgName+"-mask1-connect.bmp", mask1)
#
#

#
# kernal = cv2.getStructuringElement(cv2.MORPH_RECT, (3,3))
# cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernal, thresh)
# #
# cv2.imwrite("../images/process/0426/process/" + imgName + "-res.bmp", thresh)
# #



