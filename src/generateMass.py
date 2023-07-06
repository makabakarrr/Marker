import openpyxl
import numpy as np
import random
import os
import datetime
import cv2

from generate import Marker




dirPath = "../images/generate/{}".format(datetime.date.today().strftime("%y%m%d"))
folder = os.path.exists(dirPath)
if not folder:
        os.mkdir(dirPath)

xlsxPath = '../documents/marker_0625.xlsx'
# # 打开 XLSX 文件
wb = openpyxl.load_workbook(xlsxPath)
sheet = wb["Sheet1"]

for i in range(60, 70):
    dis_int = random.randint(0, 999)
    dis_dec = round(random.random(), 3)
    distance = dis_int + dis_dec

    angle_int = random.randint(0, 360)
    angle_dec = round(random.random(), 3)
    angle = angle_int + angle_dec

    # print(distance, angle)

    fileName = "marker_"+str(i)

    marker = Marker(str(distance), angle)
    image = marker.create()
    big_canvas = np.zeros((1000, 1000), np.uint8)
    big_canvas[199:799, 199:799] = image
    cv2.threshold(big_canvas, 2, 255, cv2.THRESH_BINARY, big_canvas)
    sheet.cell(row=i+2, column=1, value=fileName)
    sheet.cell(row=i+2, column=2, value=distance)
    sheet.cell(row=i+2, column=3, value=angle)

    savePath = dirPath + '/' + fileName + '.bmp'
    cv2.imwrite(savePath, big_canvas)

wb.save(xlsxPath)

