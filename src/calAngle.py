import openpyxl
import math

def calAngle1(point1, point2):
    angle = 0.0
    x1, y1 = point1
    x2, y2 = point2
    dx = x2 - x1
    dy = y2 - y1
    if y2 == y1:
        angle = math.pi / 2.0
        if x2 == x1:
            angle = 0.0
        elif x2 > x1:
            angle = 3.0 * math.pi / 2.0
    elif x2 > x1 and y2 < y1:
        angle = math.pi + math.atan(-dx / dy)
    elif x2 > x1 and y2 > y1:
        angle = math.pi * 2 - math.atan(dx / dy)
    elif x2 < x1 and y2 < y1:
        angle = math.pi - math.atan(dx / dy)
    elif x2 < x1 and y2 > y1:
        angle = math.atan(dx / -dy)
    elif x2<x1 and y1==y2:
        angle = math.pi
    return (angle * 180 / math.pi)


def calHandleAngle(point1, point2):
    angle = 0.0
    x1, y1 = point1
    x2, y2 = point2
    dx = x2 - x1
    dy = y2 - y1
    if x2 == x1:
        angle = math.pi / 2.0
        if y2 == y1:
            angle = 0.0
        elif y2 > x1:
            angle = 3.0 * math.pi / 2.0
    elif x2 < x1 and y2 > y1:
        angle = math.pi + math.atan(-dy / dx)
    elif x2 > x1 and y2 > y1:
        angle = math.pi * 2 - math.atan(dy / dx)
    elif x2 < x1 and y2 < y1:
        angle = math.pi - math.atan(dy / dx)
    elif x2 > x1 and y2 < y1:
        angle = math.atan(-dy / dx)
    elif x2<x1 and y1==y2:
        angle = math.pi
    return (angle * 180 / math.pi)


# filePath = "../documents/0625_angle_error.xlsx"
filePath = "../documents/mass/marker_decode_angle.xlsx"
wb = openpyxl.load_workbook(filePath)
sheet = wb['Sheet1']

# for r in range(2, 27, 2):
#     point1_x = sheet.cell(row=r, column=10).value
#     point1_y = sheet.cell(row=r, column=11).value
#     point2_x = sheet.cell(row=r+1, column=8).value
#     point2_y = sheet.cell(row=r+1, column=9).value
#     angle = calAngle1([point1_x, point1_y], [point2_x, point2_y])
#     sheet.cell(row=r, column=13, value=round(angle, 4))
#
# wb.save(filePath)

for r in range(2, 32):
    # point1_x = sheet.cell(row=r, column=9).value
    # point1_y = sheet.cell(row=r, column=10).value
    # point2_x = sheet.cell(row=r, column=11).value
    # point2_y = sheet.cell(row=r, column=12).value
    # angle = calHandleAngle([point1_x, point1_y], [point2_x, point2_y])
    # sheet.cell(row=r, column=13, value=round(angle, 4))
    angle1 = sheet.cell(row=r, column=17).value
    if angle1 < 0:
        angle1 = 360 + angle1
    # angle2 = sheet.cell(row=r, column=13).value
    sheet.cell(row=r, column=17, value=round(angle1, 4))
    # sheet.cell(row=r, column=16, value=round(abs(angle2-angle1), 4))

wb.save(filePath)

# p_angle_err = []
# sub_angle_err = []
# for i in range(2, )
