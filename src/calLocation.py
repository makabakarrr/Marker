import cv2
import openpyxl
import numpy as np
import itertools

from customFunction import calculateAffineMatrix, cvtCodePoints1, getRemainPoint

def getDistCombinations(source_combinations, source_points, dist_points):
    src_n0, src_n1, src_n2, src_n3 = source_points
    n0, n1, n2, n3 = dist_points
    dist_combinations  = []
    for s in source_combinations:
        d = []
        for point in s:
            xx, yy = point
            if [xx, yy] == [src_n0[0], src_n0[1]]:
                d.append(n0)
            elif [xx, yy] == [src_n1[0], src_n1[1]]:
                d.append(n1)
            elif [xx, yy] == [src_n2[0], src_n2[1]]:
                d.append(n2)
            else:
                d.append(n3)
        dist_combinations.append(d)

    return np.array(dist_combinations)


filePath = "../documents/0713_location_center.xlsx"
wb = openpyxl.load_workbook(filePath)
sheet = wb['Sheet1']


source_points = np.array([[10,10], [10, 30], [30, 30], [30, 10]], dtype=np.float32)
# source_points = np.array([[10,40], [40, 10], [30, 20], [30, 30]], dtype=np.float32)
# target_point = [[40, 40], [40, 10], [20, 0], [20.5, 0], [21, 0], [21.5, 0]]
# target_point = [[40, 40], [40, 10], [21.5, -0.5], [21, -0.5], [20.5, -0.5], [20, -0.5]]
# target_point = [[30, 10], [30, 20], [30, 30], [20, 10], [20, 20], [20, 30]]
# target_point =[[10, 10], [40, 40], [30, 10], [20, 10], [20, 20], [20, 30], [50, 10], [50, 20], [50, 30]]
# target_point = [[45, 10], [45, 20], [45, 30]]
# target_point = [[10, 30], [10, 20], [40, 20], [40, 30]]
# target_point = [[20, 30], [10, 20]]
# target_point = [[30, 10], [20, 10], [20, 20]]
# target_point = [[20, 21], [20.5, 21], [21, 21], [21.5, 21]]
# target_point = [[20, 21.5], [20.5, 21.5], [21, 21.5], [21.5, 21.5]]
# target_point = [[20, 22], [20.5, 22], [21, 22], [21.5, 22]]
# target_point = [[20, 22.5], [20.5, 22.5], [21, 22.5], [21.5, 22.5]]
# target_point = [[20, 18.5], [20.5, 18.5], [21, 18.5], [21.5, 18.5]]
# target_point = [[20, 17], [20.5, 17], [21, 17], [21.5, 17]]
# target_point = [[20, 10], [20.5, 10], [21, 10], [21.5, 10]]
# target_point = [[20, 9.5], [20.5, 9.5], [21, 9.5], [21.5, 9.5]]
# target_point = [[20, 21], [20.5, 21], [21, 21], [21.5, 21]]
# target_point = [[20, 20.5], [20.5, 20.5], [21, 20.5], [21.5, 20.5]]
# target_point = [[20, 20], [20.5, 20], [21, 20], [21.5, 20]]
# target_point = [[20, 8.5], [20.5, 8.5], [21, 8.5], [21.5, 8.5]]
target_point = [[20, 8], [20.5, 8], [21, 8], [21.5, 8]]
# target_real = []
dist_points = []

for row in range(2, 8):
    b = sheet.cell(row=row, column=4)
    c = sheet.cell(row=row, column=5)
    d = sheet.cell(row=row, column=6)
    e = sheet.cell(row=row, column=7)
    # sheet.cell(row=row, column=10, value=(d.value+b.value))
    # sheet.cell(row=row, column=11, value=(e.value+c.value))
    if row < 6:
        dist_points.append([d.value+b.value, e.value+c.value])
    print("坐标：", [d.value, e.value])
    # dist_points.append([d.value, e.value])

wb.save(filePath)


source_combinations = np.array(list(itertools.combinations(source_points, 3)))
dist_combinations = getDistCombinations(source_combinations, source_points, np.array(dist_points, dtype=np.float32))


angle = 0.0
transform_matrix = np.zeros((2,3), np.float32)
max_dist = 10.0

for i in range(len(source_combinations)):
    MM = calculateAffineMatrix(source_combinations[i], dist_combinations[i])
    s_remain_p, d_remain_p = getRemainPoint(source_combinations[i], source_points, dist_points)
    cal_dist_x = round(MM[0][0] * s_remain_p[0] + MM[0][1] * s_remain_p[1] + MM[0][2], 4)
    cal_dist_y = round(MM[1][0] * s_remain_p[0] + MM[1][1] * s_remain_p[1] + MM[1][2], 4)
    print(s_remain_p, cal_dist_y, cal_dist_y)
    dist_err = np.sqrt((cal_dist_x-d_remain_p[0])**2 + (cal_dist_y-d_remain_p[1])**2)
    print("dist_err:", dist_err, 135 - np.arctan2(MM[1, 0], MM[0, 0]) * 180 / np.pi)
    if dist_err<max_dist:
        max_dist = dist_err
        angle = 135 - np.arctan2(MM[1, 0], MM[0, 0]) * 180 / np.pi
        transform_matrix = MM
angle = angle if angle>0 else 360+angle
targets = cvtCodePoints1(target_point, transform_matrix)
print("targets:", targets)
print("angle", angle, max_dist)

for i in range(1, len(targets)):
    t1 = targets[i-1]
    t2 = targets[i]
    print("移动距离：", t2[0]-t1[0], t2[1]-t1[1])


